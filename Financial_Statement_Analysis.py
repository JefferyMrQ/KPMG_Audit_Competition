# -*- coding:utf-8 -*-
"""
 作者: QL
 日期: 2022年07月07日
"""
# # jupyter lab 必须在调用模块前声明
# from pyecharts.globals import CurrentConfig, NotebookType
# CurrentConfig.NOTEBOOK_TYPE = NotebookType.JUPYTER_LAB
from typing import NoReturn
from pyecharts.charts import *
from pyecharts.components import Table
from pyecharts import options as opts
from pyecharts.globals import ThemeType
from pyecharts.commons.utils import JsCode
from pyecharts.options import PageLayoutOpts
import pandas as pd
import numpy as np
import os
import webbrowser
import re
import json
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
import string
import logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(thread)d %(levelname)s %(module)s - %(message)s')
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 初始设置
# 颜色与主题设置
THEME = ThemeType.LIGHT  # 定制主题(自定义颜色时，主题失效)
COLOR_LST = ['rgb(148,162,255)', 'rgb(107,162,255)', '#A75CFB', 'rgb(255,178,255)', 'rgb(198,205,255)']  # 柱状图和折线图配色
# BACK_LINE_COLOR_LST = ['rgb(255,157,100)', 'rgb(252,114,147)', 'rgb(255,220,95)']
BACK_LINE_COLOR_LST = ['#A75CFB', '#2367F8', 'rgb(255,172,255)']  # Bar().overlap(Line())时，line的配色
PIE_COLOR_LST = ['#D896F6', '#427CFF',
                 '#EDCEF0', '#E6B2F3', '#D896F6', '#A75CFB', '#843EFE',
                 '#A6C3FD', '#72A0FE', '#568AFD', '#427CFF', '#2367F8']  # 饼状图配色
# COLOR_LST = ['#4886F4', '#5983FC', '#3E80C1', '#1F9CE4']

# 其他设置
WIDTH = 2.5  # 折现宽度设置
ANOMALOUS_BOUNDARY = 100  # 差额百分比阈值设置
os.environ['NUMEXPR_MAX_THREADS'] = '16'  # numpy设置


class DataProcess:
    def __init__(self, ticker):
        self.ticker = ticker

    def basic_info(self) -> list:
        """
        获得、处理公司基本信息
        :return: 含有公司基本信息和董监高信息的列表[基本星系Series, 董监高信息DataFrame]
        """
        path_basic = r".\\data\\" + self.ticker + r"_公司资料.csv"
        path_djg = r".\\data\\" + self.ticker + r"_董监高指标.csv"
        path_undergraduate = r".\\data\\" + self.ticker + r"_本科人数.csv"
        path_graduate = r".\\data\\" + self.ticker + r"_硕士人数.csv"
        path_doctor = r".\\data\\" + self.ticker + r"_博士人数.csv"

        df_basic = pd.read_csv(path_basic, index_col=0)
        df_djg = pd.read_csv(path_djg, index_col=0)
        df_undergraduate = pd.read_csv(path_undergraduate, index_col=0)
        df_graduate = pd.read_csv(path_graduate, index_col=0)
        df_doctor = pd.read_csv(path_doctor, index_col=0)

        s_info = df_basic.sort_index().iloc[-1, :]
        df_undergraduate.drop(['CODES'], axis=1, inplace=True)
        df_graduate.drop(['CODES'], axis=1, inplace=True)
        df_doctor.drop(['CODES'], axis=1, inplace=True)
        df_undergraduate.rename(columns={"各学历员工人数": "本科人数"}, inplace=True)
        df_graduate.rename(columns={"各学历员工人数": "硕士人数"}, inplace=True)
        df_doctor.rename(columns={"各学历员工人数": "博士人数"}, inplace=True)
        lst = [df_undergraduate, df_graduate, df_doctor]
        for df in lst:
            df_djg = pd.merge(df_djg, df, left_index=True, right_index=True, how='inner')

        return [s_info.drop('CODES'), df_djg.drop(['CODES'], axis=1).iloc[4:, :]]

    def cash_flow_statement_data(self) -> pd.DataFrame:
        """
        获得、处理现金流量表数据
        :return: 现金流量表数据
        """
        path = r".\\data\\" + self.ticker + r"_现金流量表.csv"
        df = pd.read_csv(path, index_col=0)

        return df.iloc[3:, :]

    def statement_of_profit_and_loss_data(self) -> pd.DataFrame:
        """
        获得、处理利润表数据
        :return: 利润表数据
        """
        df = pd.read_csv(r".\\data\\" + self.ticker + r"_利润表.csv", index_col=0)
        # 其他业务收入缺失值填补
        for date in list(df.index):
            if np.isnan(list(df.loc[date, ['其他业务收入(附注)']])[0]):
                df.loc[date, ['其他业务收入(附注)']] = df.loc[date, ['营业收入']].to_numpy() - df.loc[date, ['主营营业收入']].to_numpy()

        return df.iloc[3:, :]

    def statement_of_financial_position(self) -> pd.DataFrame:
        """
        获得、处理资产负债表数据
        :return: 资产负债表数据
        """
        df = pd.read_csv(r".\\data\\" + self.ticker + r"_资产负债表.csv", index_col=0)
        # 其他业务收入缺失值填补
        df.fillna(0, inplace=True)
        df['其他流动资产'] = df['流动资产合计'] - df['货币资金'] - df['应收票据及应收账款'] - df['预付款项'] - df['存货']
        df['其他非流动资产'] = df['非流动资产合计'] - df['长期股权投资'] - df['固定资产'] - df['在建工程'] - df['无形资产']
        df['其他流动负债'] = df['流动负债合计'] - df['短期借款'] - df['应付票据及应付账款'] - df['合同负债'] - df['应付职工薪酬']
        df['其他非流动负债'] = df['非流动负债合计'] - df['长期借款'] - df['应付债券'] - df['租赁负债'] - df['递延收益']
        df_final = df.iloc[4:, :]
        return df_final

    def solvency(self) -> list:
        """
        获得、处理偿债能力数据
        :return: 偿债能力DataFrame列表[企业, 行业]
        """
        path = r".\\data\\" + self.ticker + r"_偿债能力.csv"
        path_ind = r".\\data\\" + self.ticker + r"_偿债能力_行业.csv"

        df = pd.read_csv(path, index_col=0)
        df_ind = pd.read_csv(path_ind, index_col=0)

        # df.rename(columns={'经营活动产生的现金流量净额/负债合计': '经营现金流量比率'}, inplace=True)

        df_final = df.iloc[4:, :]
        df_ind_final = df_ind.iloc[4:, :]

        return [df_final, df_ind_final]

    def growth_ability(self) -> list:
        """
        获得、处理成长能力数据
        :return: 成长能力DataFrame列表[企业, 行业]
        """
        path = r".\\data\\" + self.ticker + r"_成长能力.csv"
        path_ind = r".\\data\\" + self.ticker + r"_成长能力_行业.csv"

        df = pd.read_csv(path, index_col=0)
        df_ind = pd.read_csv(path_ind, index_col=0)

        df.rename(columns={'净资产收益率同比增长率(摊薄)': '净资产收益率同比增长率'}, inplace=True)
        df_ind['总资产(同比增长率)'] = (
                (df_ind['资产总计(合计)'] - df_ind['资产总计(合计)'].shift(1)) / abs(df_ind['资产总计(合计)'].shift(1)) * 100)
        df_ind.drop(labels='资产总计(合计)', axis=1, inplace=True)

        df_final = df.iloc[4:, :]
        df_ind_final = df_ind.iloc[4:, :]

        return [df_final, df_ind_final]

    def profitability(self) -> list:
        """
        获得、处理盈利能力数据
        :return: 盈利能力DataFrame列表[企业, 行业]
        """
        path = r".\\data\\" + self.ticker + r"_盈利能力.csv"
        path_ind = r".\\data\\" + self.ticker + r"_盈利能力_行业.csv"
        path_per_stock = r".\\data\\" + self.ticker + r"_每股指标.csv"
        path_per_stock_ind = r".\\data\\" + self.ticker + r"_每股指标_行业.csv"

        df = pd.read_csv(path, index_col=0)
        df_ind = pd.read_csv(path_ind, index_col=0)
        df_per_stock = pd.read_csv(path_per_stock, index_col=0)
        df_per_stock_ind = pd.read_csv(path_per_stock_ind, index_col=0)

        df = pd.merge(df.drop(['CODES'], axis=1), df_per_stock.drop(['CODES'], axis=1), left_index=True, right_index=True, how='inner')
        df.rename(columns={'净资产收益率ROE(加权)': '净资产收益率ROE', '净利润/营业总收入': '营业净利率', '每股收益EPS(基本)': '每股收益EPS'}, inplace=True)
        df_ind = pd.merge(df_ind.drop(['CODES'], axis=1), df_per_stock_ind.drop(['CODES'], axis=1), left_index=True, right_index=True, how='inner')

        df_final = df.iloc[4:, :]
        df_ind_final = df_ind.iloc[4:, :]

        return [df_final, df_ind_final]

    def operating_capacity(self) -> list:
        """
        获得、处理营运能力数据
        :return: 营运能力DataFrame列表[企业, 行业, 应收账款明细, 存货项目明细]
        """
        path = r".\\data\\" + self.ticker + r"_营运能力.csv"
        path_ind = r".\\data\\" + self.ticker + r"_营运能力_行业.csv"
        path1 = r".\\data\\" + self.ticker + r"_应收账款明细.csv"
        path2 = r".\\data\\" + self.ticker + r"_存货项目明细.csv"

        df = pd.read_csv(path, index_col=0)
        df_ind = pd.read_csv(path_ind, index_col=0)
        df1 = pd.read_csv(path1, index_col=0)
        df2 = pd.read_csv(path2, index_col=0)

        # 统一几个比率的单位
        if (df.loc[:, '应收账款周转率(含应收票据)'] >= 100).any():
            df['应收账款周转率(含应收票据)'] = df['应收账款周转率(含应收票据)'] / 100
        if (df_ind.loc[:, '应收账款周转率(算术平均)'] >= 100).any():
            df_ind['应收账款周转率(算术平均)'] = df_ind['应收账款周转率(算术平均)'] / 100
        if (df.loc[:, '固定资产周转率'] >= 10).any():
            df['固定资产周转率'] = df['固定资产周转率'] / 100
        if (df_ind.loc[:, '固定资产周转率(算术平均)'] >= 10).any():
            df_ind['固定资产周转率(算术平均)'] = df_ind['固定资产周转率(算术平均)'] / 100

        df.rename(columns={'应收账款周转率(含应收票据)': '应收账款周转率', '应付账款周转率(含应付票据)': '应付账款周转率'}, inplace=True)

        df1['其他'] = df1['应收账款合计'] - df1['应收账款—金额'] - df1['应收账款—坏账准备']
        df1.drop(columns=['CODES', '应收账款合计', '应收账款—比例'], inplace=True)
        l1 = df1.columns.drop('其他')
        l2 = [i[1] for i in list(df1.columns.drop('其他').str.split('—'))]
        df1.rename(columns=dict(zip(l1, l2)), inplace=True)

        df2['其他'] = df2['存货明细-合计'] - df2['存货明细-原材料'] - df2['存货明细-在产品'] - df2['存货明细-产成品'] - df2['存货明细-库存商品'] - df2['存货明细-周转材料'] - df2['存货明细-委托加工材料']
        df2.drop(columns=['CODES', '存货明细-合计'], inplace=True)
        l3 = df2.columns.drop('其他')
        l4 = [i[1] for i in list(df2.columns.drop('其他').str.split('-'))]
        df2.rename(columns=dict(zip(l3, l4)), inplace=True)

        df_final = df.iloc[4:, :]
        df_ind_final = df_ind.iloc[4:, :]
        df1 = df1.iloc[4:, :]
        df2 = df2.iloc[4:, :]

        return [df_final, df_ind_final, df1, df2]

    def get_extra_trial_data(self) -> dict:
        """
        尝试将企业业绩评价标准值呈现在可视化系统中(以2012年证监会行业分类中的"天然原油和天然气开采业"行业与"中国石化"企业为例)
        :return:行业数据字典
        """
        path = r'.\\data\\trial_data\\天然原油和天然气开采业_企业业绩评价标准值.csv'
        df = pd.read_csv(path)
        data_dic = {}

        for col in df.columns[2:]:
            ts_data_dic = {}
            for level in range(1, 6):
                ts_data_lst = []
                for date in range(2017, 2022):
                    ts_data_lst.append((df.loc[(df['SgnYear'] == date) & (df['Standard'] == level), col]).values[0])
                ts_data_dic[str(level)] = ts_data_lst
            data_dic[col] = ts_data_dic

        return data_dic

    def excel_analysis1(self) -> list:
        """
        获得、处理财务分析数据1
        :return: 财务分析表格DataFrame列表[企业, 行业]
        """
        # 企业
        path1 = r".\\data\\" + self.ticker + r"_舞弊分析数据.csv"
        df1 = pd.read_csv(path1, index_col=0)

        # 数据处理
        # 增长率计算
        growth_term = "应收账款 应收票据 其他应收款合计 预付款项 存货 在建工程 长期待摊费用 固定资产 " \
                      "商誉 资产减值准备 短期借款 应付票据 应付账款 其他应付款合计 " \
                      "营业总收入 主营营业收入 营业总成本 销售毛利率 销售净利率 " \
                      "经营活动产生的现金流量净额".split(' ')
        for term in growth_term:
            df1[term + '增长率'] = (df1[term] - df1[term].shift(1)) / df1[term].shift(1)

        for term in ['应收账款', '存货', '固定资产']:
            df1[term + '比率'] = df1[term] / df1['负债和股东权益合计']
            df1[term + '比率增长率'] = (df1[term + '比率'] - df1[term + '比率'].shift(1)) / df1[term + '比率'].shift(1)

        df1['费用率'] = (df1['营业总成本'] + df1['销售费用'] + df1['管理费用'] + df1['财务费用']) / df1['营业总收入']
        df1['研发支出占比'] = df1['研发费用'] / df1['营业总成本']
        df1['带息负债和货币资金比值'] = df1['带息负债'] / df1['货币资金']
        df1['货币资金/总资产'] = df1['货币资金'] / df1['负债和股东权益合计']

        # sorted_term = [term + '增长率' for term in growth_term[: -3]] + \
        #               ['总资产周转率', '净资产收益率'] + \
        #               [term + '增长率' for term in growth_term[-3: -1]] + ['销售毛利率', '销售净利率', '费用率'] + \
        #               ['存货增长率', '固定资产增长率', '应收账款比率', '应收账款周转率', '存货比率', '存货周转率', '固定资产比率', '固定资产周转率'] + \
        #               ['研发支出占比', '经营活动产生的现金流量净额增长率', '资产负债率', '带息负债和货币资金比值']
        # return df.iloc[1:, :].loc[:, sorted_term].T

        # 行业
        path2_1 = r".\\data\\" + self.ticker + r"_舞弊分析数据_行业1.csv"
        path2_2 = r".\\data\\" + self.ticker + r"_舞弊分析数据_行业2.csv"
        df2_1 = pd.read_csv(path2_1, index_col=0)
        df2_2 = pd.read_csv(path2_2, index_col=0)
        df2 = pd.concat([df2_1, df2_2], axis=1)

        # 数据处理
        # 增长率计算
        growth_term_hy = "应收账款 预付款项 存货 固定资产 应付账款 营业总收入 营业总成本 销售毛利率 销售净利率 经营活动产生的现金流量净额".split(' ')
        for term in growth_term_hy:
            df2[term + '增长率'] = (df2[term] - df2[term].shift(1)) / df2[term].shift(1)

        for term in ['应收账款', '存货', '固定资产']:
            df2[term + '比率'] = df2[term] / df2['负债和股东权益合计']
            df2[term + '比率增长率'] = (df2[term + '比率'] - df2[term + '比率'].shift(1)) / df2[term + '比率'].shift(1)

        df2['费用率'] = (df2['营业总成本'] + df2['销售费用'] + df2['管理费用'] + df2['财务费用']) / df2['营业总收入']
        df2['货币资金/总资产'] = df2['货币资金'] / df2['负债和股东权益合计']

        df1 = df1.iloc[4:, :]
        df2 = df2.iloc[4:, :]

        return [df1, df2]

    def excel_analysis1_judge_law(self):
        """
        判断财务分析数据1
        :return: 财务分析判断结果表格DataFrame
        """ 
        df1 = self.excel_analysis1()[0].copy()  # 企业数据
        df2 = self.excel_analysis1()[1].copy()  # 行业数据

        G = '高'
        Z = '中'
        D = '低'

        # 企业自身比较
        def law1(s, upper, intermediate: list):
            if s >= upper:
                return G
            elif intermediate[0] <= s < intermediate[1]:
                return Z
            elif not pd.isnull(s):
                return D

        # 资产类
        # 应收账款增长率
        df1['qy应收账款增长率'] = df1['应收账款增长率'].apply(law1, args=(1.5, [0.6, 1.5]))

        # 应收票据增长率
        df1['qy应收票据增长率'] = df1['应收票据增长率'].apply(law1, args=(10, [2, 10]))

        # 其他应收款合计增长率
        df1['qy其他应收款合计增长率'] = df1['其他应收款合计增长率'].apply(law1, args=(10, [1.8, 10]))

        # 预付款项增长率
        df1['qy预付款项增长率'] = df1['预付款项增长率'].apply(law1, args=(5, [1.5, 5]))

        # 存货增长率
        df1['qy存货增长率'] = df1['存货增长率'].apply(law1, args=(2, [0.6, 2]))

        # 在建工程增长率
        df1['qy在建工程增长率'] = df1['在建工程增长率'].apply(law1, args=(30, [5, 30]))

        # 长期待摊费用增长率
        df1['qy长期待摊费用增长率'] = df1['长期待摊费用增长率'].apply(law1, args=(10, [2, 10]))

        # 固定资产增长率
        df1['qy固定资产增长率'] = df1['固定资产增长率'].apply(law1, args=(2, [0.6, 2]))

        # 商誉增长率
        df1['qy商誉增长率'] = df1['商誉增长率'].apply(law1, args=(30, [5, 30]))

        # 资产减值准备增长率
        df1['qy资产减值准备增长率'] = df1['资产减值准备增长率'].apply(law1, args=(30, [5, 30]))

        # 负债类
        # 短期借款增长率
        df1['qy短期借款增长率'] = df1['短期借款增长率'].apply(law1, args=(6, [1.5, 6]))

        # 应付票据增长率
        df1['qy应付票据增长率'] = df1['应付票据增长率'].apply(law1, args=(7.5, [2, 7.5]))

        # 应付账款增长率
        df1['qy应付账款增长率'] = df1['应付账款增长率'].apply(law1, args=(1.7, [0.7, 1.7]))

        # 其他应付款合计增长率
        df1['qy其他应付款合计增长率'] = df1['其他应付款合计增长率'].apply(law1, args=(10, [1.6, 10]))

        # 资产带来收入
        # 营业总收入增长率
        df1['qy营业总收入增长率'] = df1['营业总收入增长率'].apply(law1, args=(1, [0.5, 1]))

        # 主营营业收入增长率
        df1['qy主营营业收入增长率'] = df1['主营营业收入增长率'].apply(law1, args=(1, [0.5, 1]))

        # 营业总成本增长率
        df1['qy营业总成本增长率'] = df1['营业总成本增长率'].apply(law1, args=(1, [0.5, 1]))

        # 收入带来利润
        # 销售毛利率增长率
        df1['qy销售毛利率增长率'] = df1['销售毛利率增长率'].apply(law1, args=(0.8, [0.2, 0.8]))

        # 销售净利率增长率
        df1['qy销售净利率增长率'] = df1['销售净利率增长率'].apply(law1, args=(3, [1, 3]))

        # 利润带来现金流
        # 存货增长率
        # df1['存货增长率'] = df1['存货增长率'].apply(law1, args=(2, [0.6, 2]))

        # 固定资产增长率
        # df1['固定资产增长率'] = df1['固定资产增长率'].apply(law1, args=(2, [0.6, 2]))

        # 应收账款比率增长率
        df1['qy应收账款比率增长率'] = df1['应收账款比率增长率'].apply(law1, args=(1.5, [0.5, 1.5]))

        # 存货比率增长率
        df1['qy存货比率增长率'] = df1['存货比率增长率'].apply(law1, args=(1, [0.4, 1]))

        # 固定资产比率增长率
        df1['qy固定资产比率增长率'] = df1['固定资产比率增长率'].apply(law1, args=(1.8, [0.5, 1.8]))

        # 经营活动产生的现金流量净额增长率
        df1['qy经营活动产生的现金流量净额增长率'] = df1['经营活动产生的现金流量净额增长率'].apply(law1, args=(10, [2, 10]))

        # 企业与行业比较
        def law2(s, type: str):
            if type == 'pos':
                if s >= 1:
                    return G
                elif 0.25 <= s < 1:
                    return Z
                elif not pd.isnull(s):
                    return D
            elif type == 'neg':
                if s <= -1:
                    return G
                elif -1 < s <= -0.25:
                    return Z
                elif not pd.isnull(s):
                    return D
            elif type == 'abs':
                if abs(s) > 1:
                    return G
                elif 0.25 <= abs(s) < 1:
                    return Z
                elif not pd.isnull(abs(s)):
                    return D

        def do_template2(df1, df2, feature_name: str, type: str):
            df1['vs' + feature_name] = (df1[feature_name] - df2[feature_name]) / df2[feature_name]
            df1['vs' + feature_name] = df1['vs' + feature_name].apply(law2, args=(type, ))

        # 资产类
        # 预付款项增长率
        do_template2(df1, df2, '预付款项增长率', 'pos')

        # 应收账款增长率
        do_template2(df1, df2, '应收账款增长率', 'pos')

        # 存货增长率
        do_template2(df1, df2, '存货增长率', 'pos')

        # 固定资产增长率
        do_template2(df1, df2, '固定资产增长率', 'pos')

        # 负债类
        # 应付账款增长率
        do_template2(df1, df2, '应付账款增长率', 'pos')

        # 资产带来收入
        # 总资产周转率
        do_template2(df1, df2, '总资产周转率', 'neg')

        # 净资产收益率
        do_template2(df1, df2, '净资产收益率', 'abs')

        # 收入带来利润
        # 销售毛利率
        do_template2(df1, df2, '销售毛利率', 'abs')

        # 销售净利率
        do_template2(df1, df2, '销售净利率', 'abs')

        # 销售费用率(费用率)
        do_template2(df1, df2, '费用率', 'pos')

        # 利润带来现金流
        # 应收账款周转率
        do_template2(df1, df2, '应收账款周转率', 'neg')

        # 存货周转率
        do_template2(df1, df2, '存货周转率', 'neg')

        # 固定资产周转率
        do_template2(df1, df2, '固定资产周转率', 'neg')

        # 应收账款比率增长率
        do_template2(df1, df2, '应收账款比率增长率', 'pos')

        # 存货比率增长率
        do_template2(df1, df2, '存货比率增长率', 'pos')

        # 固定资产比率增长率
        do_template2(df1, df2, '固定资产比率增长率', 'pos')

        # 存贷双高
        # 货币资金/总资产
        do_template2(df1, df2, '货币资金/总资产', 'pos')

        # 资产负债率
        do_template2(df1, df2, '资产负债率', 'pos')

        return df1

    def excel_analysis2(self) -> list:
        """
        获得、处理财务分析数据2
        :return: 财务分析表格DataFrame
        """
        # 企业数据
        path = r"./data/" + (self.ticker).strip() + r"_舞弊分析数据.csv"
        df = pd.read_csv(path, index_col=0)

        # 行业数据
        path2_1 = r".\\data\\" + self.ticker + r"_舞弊分析数据_行业1.csv"
        path2_2 = r".\\data\\" + self.ticker + r"_舞弊分析数据_行业2.csv"
        df2_1 = pd.read_csv(path2_1, index_col=0)
        df2_2 = pd.read_csv(path2_2, index_col=0)
        df2 = pd.concat([df2_1, df2_2], axis=1)

        # 数据计算
        def groth_rate(df, col):
            return (df[col] - df[col].shift(1)) / df[col].shift(1)

        # 固定资产占总资产的比重变化率
        df['固定资产占总资产的比重'] = df['固定资产'] / df['负债和股东权益合计']
        df['固定资产占总资产的比重变化率'] = groth_rate(df, '固定资产占总资产的比重')

        # 固定资产增加率
        df['固定资产增加率'] = groth_rate(df, '固定资产')
        # 累计折旧占固定资产原值的比重变化率
        df['累计折旧占固定资产原值的比重'] = df['固定资产累计折旧'] / df['固定资产']
        df['累计折旧占固定资产原值的比重变化率'] = groth_rate(df, '累计折旧占固定资产原值的比重')

        # 存货占总资产比重变化率
        df['存货占总资产比重'] = df['存货'] / df['负债和股东权益合计']
        df['存货占总资产比重变化率'] = groth_rate(df, '存货占总资产比重')

        # 存货增加率
        df['存货增加率'] = groth_rate(df, '存货')
        # 营业收入增长率
        df['营业收入增长率'] = groth_rate(df, '营业总收入')
        # 营业成本增长率
        df['营业成本增长率'] = groth_rate(df, '营业总成本')

        # 存货跌价准备占存货的比重变化率
        df['存货跌价准备占存货的比重'] = df['存货跌价准备合计'] / df['应收账款合计']
        df['存货跌价准备占存货的比重变化率'] = groth_rate(df, '存货跌价准备占存货的比重')

        # 应收账款增长率
        df['应收账款增长率'] = groth_rate(df, '应收账款')
        # 存货增长率
        df['存货增长率'] = groth_rate(df, '存货')
        # 营业收入增长率
        df['营业收入增长率'] = groth_rate(df, '营业总收入')

        # 现金及现金等价物占总资产的比重变化率
        df['现金及现金等价物占总资产的比重'] = df['期末现金及现金等价物余额'] / df['负债和股东权益合计']
        df['现金及现金等价物占总资产的比重变化率'] = groth_rate(df, '现金及现金等价物占总资产的比重')
        # 应收账款占总资产的比重增长率
        df['应收账款占总资产的比重'] = df['应收账款'] / df['负债和股东权益合计']
        df['应收账款占总资产的比重增长率'] = groth_rate(df, '应收账款占总资产的比重')

        # 存货占总资产比重变化率
        df['存货占总资产比重'] = df['存货'] / df['负债和股东权益合计']
        df['存货占总资产比重变化率'] = groth_rate(df, '存货占总资产比重')

        # 坏账占应收账款比重的变化率
        df['坏账占应收账款比重'] = df['坏账准备合计'] / df['应收账款合计']
        df['坏账占应收账款比重变化率'] = groth_rate(df, '坏账占应收账款比重')

        # 销售费用增长速度
        df['销售费用增长速度'] = groth_rate(df, '销售费用')
        # 营业收入增长速度
        df['营业收入增长速度'] = groth_rate(df, '营业总收入')

        # 研发费用增长率
        df['研发费用增长率'] = groth_rate(df, '研发费用')
        # 营业收入增长率
        df['营业收入增长率'] = groth_rate(df, '营业总收入')

        # 毛利率变化率
        df['毛利率变化率'] = groth_rate(df, '销售毛利率')

        # 毛利率变化率
        # 存货周转率变化率
        df['存货周转率变化率'] = groth_rate(df, '存货周转率')

        # 毛利率变化率
        # 应收账款周转率变化率
        df['应收账款周转率变化率'] = groth_rate(df, '应收账款周转率')

        # 毛利率变化率
        # 应付账款变化率
        df['应付账款变化率'] = groth_rate(df, '应付账款')

        # 毛利率
        # 现金循环周期
        df['现金循环周期'] = df['存货周转天数'] + df['应收账款周转天数'] - df['应付账款周转天数']
        # 行业毛利率
        df['行业毛利率'] = df2['销售毛利率']
        # 行业现金循环周期
        df['行业现金循环周期'] = df2['存货周转天数'] + df2['应收账款周转天数'] - df2['应付账款周转天数']

        # 毛利率
        # 销售费用率
        df['销售费用率'] = (df['营业总成本'] + df['销售费用'] + df['管理费用'] + df['财务费用']) / df['营业总收入']
        # 行业毛利率
        df['行业毛利率'] = df2['销售毛利率']
        # 行业销售费用率
        df['行业销售费用率'] = (df2['营业总成本'] + df2['销售费用'] + df2['管理费用'] + df2['财务费用']) / df2['营业总收入']

        # return df.iloc[4:, :]
        return df

    def excel_analysis2_judge_law(self):
        """
        判断财务分析数据2
        :return: 财务分析判断结果表格DataFrame
        """
        df = self.excel_analysis2().copy()
        G = '高'
        Z = '中'
        D = '低'

        # 一、购货与付款循环重大错报风险的识别
        # （1）固定资产异常：固定资产占总资产的比重增长异常
        def law11(s):
            if s >= 2:
                return G
            elif 0.5 <= s < 2:
                return Z
            elif not pd.isnull(s):
                return D

        df['law11'] = df['固定资产占总资产的比重变化率'].map(law11)

        # （2）固定资产异常：固定资产与累计折旧占固定资产原值比重联动异常
        df['law12'] = df['固定资产增加率'] - df['累计折旧占固定资产原值的比重变化率']
        for row_num in range(df.shape[0]):
            row = df.index[row_num]
            if (df.loc[row, '固定资产增加率'] > 0) and (df.loc[row, 'law12'] >= 3):
                df.loc[row, 'law12'] = G
            elif (df.loc[row, '固定资产增加率'] > 0) and (1 <= df.loc[row, 'law12'] < 3):
                df.loc[row, 'law12'] = Z
            elif not pd.isnull(df.loc[row, 'law12']):
                df.loc[row, 'law12'] = D

        # 二、生产与薪酬循环重大错报风险的识别
        # （1）存货异常：存货占总资产的比重增长异常
        def law21(s):
            if s >= 1:
                return G
            elif 0.5 < s < 1:
                return Z
            elif not pd.isnull(s):
                return D

        df['law21'] = df['存货占总资产比重变化率'].map(law21)

        # （2）存货异常：存货与营业收入联动异常
        df['law22'] = df['存货增长率'] - df['营业收入增长率']
        for row_num in range(df.shape[0]):
            row = df.index[row_num]
            if (df.loc[row, '存货增长率'] > 0) and (df.loc[row, 'law22'] >= 1.2):
                df.loc[row, 'law22'] = G
            elif (df.loc[row, '存货增长率'] > 0) and (0.5 <= df.loc[row, 'law22'] < 1.2):
                df.loc[row, 'law22'] = Z
            elif not pd.isnull(df.loc[row, 'law22']):
                df.loc[row, 'law22'] = D

        # （3）存货异常：存货与营业成本联动异常
        df['law23'] = df['存货增长率'] - df['营业成本增长率']
        for row_num in range(df.shape[0]):
            row = df.index[row_num]
            if (df.loc[row, '存货增长率'] > 0) and (df.loc[row, 'law23'] >= 1.2):
                df.loc[row, 'law23'] = G
            elif (df.loc[row, '存货增长率'] > 0) and (0.5 <= df.loc[row, 'law23'] < 1.2):
                df.loc[row, 'law23'] = Z
            elif not pd.isnull(df.loc[row, 'law23']):
                df.loc[row, 'law23'] = D

        # （4）存货异常：存货跌价准备占存货比重异常
        def law24(s):
            if (s >= 12) or (s <= -0.9):
                return G
            elif (2 <= s < 12) or (-0.9 < s <= -0.6):
                return Z
            elif not pd.isnull(s):
                return D

        df['law24'] = df['存货跌价准备占存货的比重变化率'].map(law24)

        # （5）应付账款异常：应付账款与存货联动异常
        df['law25'] = df['应付账款变化率'] - df['存货增长率']
        for row_num in range(df.shape[0]):
            row = df.index[row_num]
            if (df.loc[row, '应付账款变化率'] > 0) and (df.loc[row, 'law25'] >= 1.2):
                df.loc[row, 'law25'] = G
            elif (df.loc[row, '应付账款变化率'] > 0) and (0.5 <= df.loc[row, 'law25'] < 1.2):
                df.loc[row, 'law25'] = Z
            elif not pd.isnull(df.loc[row, 'law25']):
                df.loc[row, 'law25'] = D

        # （6）应付账款异常：应付账款与营业收入联动异常
        df['law26'] = df['应付账款变化率'] - df['营业收入增长率']
        for row_num in range(df.shape[0]):
            row = df.index[row_num]
            if (df.loc[row, '应付账款变化率'] > 0) and (df.loc[row, 'law26'] >= 1.5):
                df.loc[row, 'law26'] = G
            elif (df.loc[row, '应付账款变化率'] > 0) and (0.6 <= df.loc[row, 'law26'] < 1.5):
                df.loc[row, 'law26'] = Z
            elif not pd.isnull(df.loc[row, 'law26']):
                df.loc[row, 'law26'] = D

        # 三、销售与收款循环重大错报风险的识别
        # （1）应收账款异常：应收账款与现金及现金等价物联动异常
        df['law31'] = df['应收账款增长率'] - df['现金及现金等价物占总资产的比重变化率']
        for row_num in range(df.shape[0]):
            row = df.index[row_num]
            if (df.loc[row, '应收账款增长率'] > 1.5) and (df.loc[row, '现金及现金等价物占总资产的比重变化率'] < 0):
                df.loc[row, 'law31'] = G
            elif (0.5 < df.loc[row, '应收账款增长率'] < 1.5) and (df.loc[row, '现金及现金等价物占总资产的比重变化率'] < 0):
                df.loc[row, 'law31'] = Z
            elif not pd.isnull(df.loc[row, 'law31']):
                df.loc[row, 'law31'] = D

        # （2）应收账款异常：坏账准备占应收账款比重联动异常
        def law32(s):
            if s >= 4:
                return G
            elif 0.8 <= s < 4:
                return Z
            elif not pd.isnull(s):
                return D

        df['law32'] = df['坏账占应收账款比重变化率'].map(law32)

        # （3）销售费用异常:销售费用与营业收入联动异常
        def law33(s):
            if s <= -1:
                return G
            elif -0.4 <= s < -1:
                return Z
            elif not pd.isnull(s):
                return D

        df['law33'] = df['销售费用增长速度'] - df['营业收入增长率']
        df['law33'] = df['law33'].map(law33)

        # （4）研发费用异常:研发费用与营业收入联动异常
        def law34(s):
            if s <= -1:
                return G
            elif -0.4 <= s < -1:
                return Z
            elif not pd.isnull(s):
                return D

        df['law34'] = df['研发费用增长率'] - df['营业收入增长率']
        df['law34'] = df['law34'].map(law34)

        # （5）毛利率异常：毛利率变动量异常
        def law35(s):
            if s >= 0.5:
                return G
            elif 0.1 <= s < 0.5:
                return Z
            elif not pd.isnull(s):
                return D

        df['law35'] = df['毛利率变化率'].map(law35)

        # （6）毛利率异常：毛利率与存货周转率联动异常
        df['law36'] = np.nan
        for row_num in range(4, df.shape[0]):
            row = df.index[row_num]
            count_lst = []  # 记录历史五年数据是否异常
            for i in range(5):
                sub_row = df.index[row_num - i]
                if (df.loc[sub_row, '毛利率变化率'] > 0) and (df.loc[sub_row, '存货周转率变化率'] < 0):
                    count_lst.append(1)
                else:
                    count_lst.append(0)
            if sum(count_lst) >= 4:
                df.loc[row, 'law36'] = G
            elif 2 <= sum(count_lst) <= 3:
                df.loc[row, 'law36'] = Z
            else:
                df.loc[row, 'law36'] = D

        # （7）毛利率异常：毛利率与应收账款周转率联动异常
        df['law37'] = np.nan
        for row_num in range(4, df.shape[0]):
            row = df.index[row_num]
            count_lst = []  # 记录历史五年数据是否异常
            for i in range(5):
                sub_row = df.index[row_num - i]
                if (df.loc[sub_row, '毛利率变化率'] > 0) and (df.loc[sub_row, '应收账款周转率变化率'] < 0):
                    count_lst.append(1)
                else:
                    count_lst.append(0)
            if sum(count_lst) >= 4:
                df.loc[row, 'law37'] = G
            elif 2 <= sum(count_lst) <= 3:
                df.loc[row, 'law37'] = Z
            else:
                df.loc[row, 'law37'] = D

        # （8）毛利率异常：毛利率与应付账款联动异常
        df['law38'] = np.nan
        for row_num in range(4, df.shape[0]):
            row = df.index[row_num]
            count_lst = []  # 记录历史五年数据是否异常
            for i in range(5):
                sub_row = df.index[row_num - i]
                if (df.loc[sub_row, '毛利率变化率'] > 0) and (df.loc[sub_row, '应付账款变化率'] < 0):
                    count_lst.append(1)
                else:
                    count_lst.append(0)
            if sum(count_lst) >= 4:
                df.loc[row, 'law38'] = G
            elif 2 <= sum(count_lst) <= 3:
                df.loc[row, 'law38'] = Z
            else:
                df.loc[row, 'law38'] = D

        # （9）毛利率异常：毛利率与现金循环周期联动异常
        df['law39'] = (df['销售毛利率'] - df['行业毛利率']) / df['行业毛利率'] + (df['现金循环周期'] - df['行业现金循环周期']) / df['行业现金循环周期']

        def law39(s):
            if s >= 1:
                return G
            elif 0.3 <= s < 1:
                return Z
            elif not pd.isnull(s):
                return D

        df['law39'] = df['law39'].map(law39)

        # （10）毛利率异常：毛利率与销售费用率联动异常
        df['law310'] = (df['销售毛利率'] - df['行业毛利率']) / df['行业毛利率'] + (df['销售费用率'] - df['行业销售费用率']) / df['行业销售费用率']

        def law310(s):
            if s >= 1:
                return G
            elif 0.3 <= s < 1:
                return Z
            elif not pd.isnull(s):
                return D

        df['law310'] = df['law310'].map(law310)

        return df.iloc[4:, :]


class BasicInfo:
    def __init__(self, ticker):
        self.ticker = ticker

    def get_info(self) -> list:
        """
        企业基本信息表格
        :return:[Table1, Table2]
        """
        s = DataProcess(self.ticker).basic_info()[0]

        # Table1
        headers1 = ['公司中文名称', '股票代码', '首发上市日期', '成立日期', '注册资本', '省份', '城市', '所属中信行业(2020)']
        row1 = list(s[headers1])
        table1 = Table().add(headers=headers1, rows=[row1])

        # Table2
        headers2 = ['主营业务']
        row2 = list(s[headers2])
        table2 = Table().add(headers=headers2, rows=[row2])

        # Table3
        headers3 = ['审计机构', '所属证监局', '律师事务所', '证券事务代表']
        row3 = list(s[headers3])
        table3 = Table().add(headers=headers3, rows=[row3])

        return [table1, table2, table3]

    def get_staff_and_salary(self) -> dict:
        """
        员工信息和薪资水平信息的八张图
        :return: 含Line的字典
        """

        df = DataProcess(self.ticker).basic_info()[1]
        x = [i[:4] for i in list(df.index)]

        c_dic = {}

        for name in df.columns[:-3]:

            if name == '员工总数':
                y = [round(i, 2) for i in list(df[name])]
                subtitle = '单位: 人'
                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>' + params[i].data[1] + '人' + '<br />';
                    }
                    return form;
                    }
                    """
                )
            else:
                y = [round(i / 10000, 2) for i in list(df[name])]
                subtitle = '单位: 万元'
                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>' + params[i].data[1] + '万元' + '<br />';
                    }
                    return form;
                    }
                    """
                )

            line = (
                Line(init_opts=opts.InitOpts(chart_id=name))
                    .add_xaxis(xaxis_data=x)
                    .add_yaxis(series_name='',
                               y_axis=y,
                               color=COLOR_LST[1],
                               is_connect_nones=True)
                    .set_series_opts(areastyle_opts=opts.AreaStyleOpts(opacity=0.65),
                                     linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title=name, subtitle=subtitle),
                    datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                    xaxis_opts=opts.AxisOpts(name='年份',
                                             name_location='middle',
                                             name_gap=25,
                                             boundary_gap=False),
                    yaxis_opts=opts.AxisOpts(name='数\n值',
                                             name_location='middle',
                                             name_gap=50,
                                             name_rotate=0,
                                             splitline_opts=opts.SplitLineOpts(is_show=True)),
                    tooltip_opts=opts.TooltipOpts(trigger='axis', formatter=formatter_line, axis_pointer_type='cross'),
                    toolbox_opts=opts.ToolboxOpts(
                        feature=opts.ToolBoxFeatureOpts(
                            data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                            brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                            magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                        )
                    )
                )
            )

            if name == '员工总数':
                pass
            else:
                line.set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                               position='top',
                                                               formatter=JsCode(
                                                                   "function(x){return x.data[1] + '万';}"
                                                               ))
                                     )

            c_dic[name] = line

        return c_dic

    def get_staff_edu(self) -> Bar:
        """
        公司员工学历比例柱状图
        :return: Bar.overlap.Pie
        """
        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '人' + '<br />';
            }
            return form;
            }
            """
        )

        df = DataProcess(self.ticker).basic_info()[1].iloc[:, -3:]

        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='staff_edu'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(df.index)])
                .add_yaxis(series_name='本科',
                           y_axis=list(df.iloc[:, 0]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.85),
                           category_gap='40%'
                           )
                .add_yaxis(series_name='硕士',
                           y_axis=list(df.iloc[:, 1]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], opacity=0.85),
                           category_gap='40%'
                           )
                .add_yaxis(series_name='博士',
                           y_axis=list(df.iloc[:, 2]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2], opacity=0.85),
                           category_gap='40%'
                           )
                .set_global_opts(
                title_opts=opts.TitleOpts(title='各学历员工人数', subtitle='单位: 人'),
                datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                xaxis_opts=opts.AxisOpts(name='年份', name_location='middle', name_gap=25),
                yaxis_opts=opts.AxisOpts(name='人\n数',
                                         name_location='middle',
                                         name_gap=50,
                                         name_rotate=0,
                                         splitline_opts=opts.SplitLineOpts(is_show=True)),
                tooltip_opts=opts.TooltipOpts(trigger='axis', formatter=formatter_bar, axis_pointer_type='none'),
                toolbox_opts=opts.ToolboxOpts(
                    feature=opts.ToolBoxFeatureOpts(
                        data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                        brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                        magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['line'])
                    )
                )
            )
        )

        return bar

    @staticmethod
    def paste_id() -> NoReturn:
        """
        寻找html中Table的id（因不是标准echarts图表，所以无法指定id）
        :return: NoReturn
        """
        soup = BeautifulSoup(open(r'.\data\raw_figure_html\公司资料_raw.html', encoding='utf-8'), features="lxml")
        div_lst = soup.find_all('div', class_='chart-container')
        id_lst = []

        # 获得html的id列表
        for i in div_lst:
            id = re.search("(?<=id=\").+?(?=\")", str(i)).group()
            id_lst.append(id)
        # 寻找Table的id
        table_id_lst = []
        for id in id_lst:
            if len(id) > 10:
                table_id_lst.append(id)
            else:
                pass
        # 替换Table的id
        # 读
        with open(r'.\data\config_data\chart_config_BasicInfo.json', 'rb') as r:
            count = 0
            params = json.load(r)
            for dic in params:
                if len(dic['cid']) > 10:
                    dic['cid'] = table_id_lst[count]
                    count += 1
                else:
                    pass
        r.close()
        # 写
        with open(r'.\data\config_data\chart_config_BasicInfo.json', 'w') as w:
            json.dump(params, w)
        w.close()

    def plot(self):
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="公司资料", layout=PageLayoutOpts(margin='10%;'))
        df = DataProcess(self.ticker).basic_info()[1]
        page.add(self.get_info()[0]).add(self.get_info()[1]).add(self.get_info()[2])
        for name in df.columns[:-3]:
            page.add(self.get_staff_and_salary()[name])
        page.add(self.get_staff_edu())
        raw_path = r'.\data\raw_figure_html\公司资料_raw.html'
        cfg_path = r'.\data\config_data\chart_config_BasicInfo.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_公司资料.html'
        page.render(raw_path)
        BasicInfo(self.ticker).paste_id()
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class CashFlowStatement:
    def __init__(self, ticker):
        self.ticker = ticker

    def get_netflow(self) -> list:
        """
        现金流量净额时间趋势图
        :return: 含有四个现金流净额的列表
        """
        df = DataProcess(self.ticker).cash_flow_statement_data()
        df0 = df[['经营活动产生的现金流量净额', '投资活动产生的现金流量净额', '筹资活动产生的现金流量净额', '现金及现金等价物净增加额']]
        x = df0.index[- (len(df0.index) - 1):]
        data_dic = df0.to_dict(orient='series')
        data = {}
        for key, val in data_dic.items():
            num = (val / 1e8).round(3)
            yoy = list(((val - val.shift(1)) / val.shift(1) * 100).round(2))  # 同比增长率
            data[key] = [num, yoy]

        formatter_bar = JsCode(  # formatter为标签内容格式器{a}：系列名;{b}：数据名;{c}：数值数组也可以是回调函数
            """function(params) {
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = color_icon + params.color + '"></span>' + params.name + '年:' + params.data + '亿元';
            return form;
        }"""
        )
        formatter_line1 = JsCode(
            """function(params) {
            form = params.value[1] + '%';
            return form;
        }"""
        )
        formatter_line2 = JsCode(
            """function(params) {
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = color_icon + params.color + '"></span>' + params.name + '年:' + params.data[1] + '%';
            return form;
        }"""
        )

        # chart_id
        chart_id_lst = ['netflow0', 'netflow1', 'netflow2', 'netflow3']
        count = 0

        # 图片列表
        c_list = []
        for key, val in data.items():
            bar = (
                Bar(init_opts=opts.InitOpts(theme=THEME, chart_id=chart_id_lst[count]))
                    .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                    .add_yaxis(
                    series_name='现金流量',
                    y_axis=list(val[0][- (len(val[0]) - 1):]),
                    itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.75),
                    category_gap='40%',
                    tooltip_opts=opts.TooltipOpts(formatter=formatter_bar)
                )
                    .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                               position='top',
                                                               font_size=10,
                                                               formatter=JsCode(
                                                                   "function(params){return Number(params.data).toFixed(1) + '亿';}"
                                                               ))
                                     )
                    .extend_axis(
                    yaxis=opts.AxisOpts(name=' 同\n 比\n 增\n 长\n 率\n(%)',
                                        name_location='middle',
                                        name_gap=50,
                                        name_rotate=0,
                                        axislabel_opts=opts.LabelOpts(formatter='{value}%')))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title=key, subtitle='单位: 元'),
                    datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                    xaxis_opts=opts.AxisOpts(name='年份', name_location='middle', name_gap=25),
                    yaxis_opts=opts.AxisOpts(name='现 \n金 \n流 \n量 ',
                                             name_location='middle',
                                             name_gap=40,
                                             name_rotate=0,
                                             splitline_opts=opts.SplitLineOpts(is_show=True)),
                )
            )
            line = (
                Line()
                    .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                    .add_yaxis(
                    series_name='同比增长率',
                    y_axis=val[1][- (len(val[1]) - 1):],
                    yaxis_index=1,
                    label_opts=opts.LabelOpts(formatter=formatter_line1),
                    tooltip_opts=opts.TooltipOpts(formatter=formatter_line2),
                    is_connect_nones=True,
                    itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[3])
                )
                    .set_series_opts(linestyle_opts=opts.LineStyleOpts(width=WIDTH))
            )
            count += 1
            bar.overlap(line)

            c_list.append(bar)
        return c_list

    def get_sankey(self) -> Timeline:
        """
        :return: 现金流桑基图
        """
        df = DataProcess(self.ticker).cash_flow_statement_data()
        df = df[- (len(df) - 1):]

        tl = Timeline(init_opts=opts.InitOpts(chart_id='sankey_timeline'))

        for i in df.index:
            nodes = [
                {"name": "经营活动流入"},
                {"name": "筹资活动流入"},
                {"name": "投资活动流入"},
                {"name": "经营活动流出"},
                {"name": "筹资活动流出"},
                {"name": "投资活动流出"},
                {"name": " "},
            ]
            links = [
                {"source": "经营活动流入", "target": " ", "value": round(float(df.loc[i, ['经营活动现金流入小计']] / 1e8), 3)},
                {"source": "投资活动流入", "target": " ", "value": round(float(df.loc[i, ['投资活动现金流入小计']] / 1e8), 3)},
                {"source": "筹资活动流入", "target": " ", "value": round(float(df.loc[i, ['筹资活动现金流入小计']] / 1e8), 3)},
                {"source": " ", "target": "经营活动流出", "value": round(float(df.loc[i, ['经营活动现金流出小计']] / 1e8), 3)},
                {"source": " ", "target": "投资活动流出", "value": round(float(df.loc[i, ['投资活动现金流入小计']] / 1e8), 3)},
                {"source": " ", "target": "筹资活动流出", "value": round(float(df.loc[i, ['筹资活动现金流入小计']] / 1e8), 3)},
            ]
            c = (
                Sankey()  # sankey图无法设置chart_id
                    .add(
                    series_name="",
                    nodes=nodes,
                    links=links,
                    pos_top="10%",
                    pos_bottom="10%",
                    label_opts=opts.LabelOpts(position="right"),
                    levels=[
                        opts.SankeyLevelsOpts(
                            depth=0,
                            itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], border_color="#f0f0f0",
                                                              border_width=0.1),
                            linestyle_opts=opts.LineStyleOpts(color="source", opacity=0.3, curve=0.5),
                        ),
                        opts.SankeyLevelsOpts(
                            depth=1,
                            itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[4], border_color="#f0f0f0",
                                                              border_width=0.1),
                            linestyle_opts=opts.LineStyleOpts(color="source", opacity=0.3, curve=0.5),
                        ),
                        opts.SankeyLevelsOpts(
                            depth=2,
                            itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], border_color="#f0f0f0",
                                                              border_width=0.1),
                            linestyle_opts=opts.LineStyleOpts(color="source", opacity=0.3, curve=0.5),
                        )
                    ]
                )
                    .set_global_opts(title_opts=opts.TitleOpts(title="现金流桑基图", subtitle="单位:亿元"))
            )
            tl.add(c, f"{i[:4]}年").add_schema(play_interval=1000, is_loop_play=False, pos_bottom="-5px", pos_left="0px")

        return tl

    def get_inflow(self) -> Bar:
        """
        :return: 现金流入分析图(Bar_Overlap_Line)
        """
        df = DataProcess(self.ticker).cash_flow_statement_data()

        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '%' + '<br />';
            }
            return form;
            }
            """
        )

        formatter_line = JsCode(
            """function(params) {
            form =
            params.name + '年' + '<br />'
            + '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:'
            + params.color + '"></span>'
            + params.seriesName + '：'
            + params.data[1] + '%';

            return form;
        }"""
        )

        formatter_line_label = JsCode(
            """function(params) {
            form = params.value[1] + '%';
            return form;
        }"""
        )

        x = df.index[- (len(df.index) - 1):]
        data1 = df['经营活动现金流入小计']
        data2 = df['投资活动现金流入小计']
        data3 = df['筹资活动现金流入小计']
        tot = data1 + data2 + data3
        d1 = (data1 / tot * 100).round(2)
        d2 = (data2 / tot * 100).round(2)
        d3 = (data3 / tot * 100).round(2)
        yoy1 = list(((d1 - d1.shift(1)) / d1.shift(1) * 100).round(2))  # 占比同比增长率
        yoy2 = list(((d2 - d2.shift(1)) / d2.shift(1) * 100).round(2))
        yoy3 = list(((d3 - d3.shift(1)) / d3.shift(1) * 100).round(2))

        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='inflow'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='经营活动现金流入占比',
                           y_axis=list(d1[- (len(d1) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.75),
                           gap=0)
                .add_yaxis(series_name='投资活动现金流入占比',
                           y_axis=list(d2[- (len(d2) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2], opacity=0.75),
                           gap=0)
                .add_yaxis(series_name='筹资活动现金流入占比',
                           y_axis=list(d3[- (len(d3) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], opacity=0.75),
                           gap=0)
                .extend_axis(
                yaxis=opts.AxisOpts(name=' 现\n 金\n 流\n 入\n 同\n 比\n 增\n 长\n 率\n(%)', name_location='middle', name_gap=50,
                                    name_rotate=0, axislabel_opts=opts.LabelOpts(formatter='{value}%')))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=JsCode(
                                                               "function(params){return Number(params.data).toFixed(1) + '%';}"
                                                           ))
                                 )
                .set_global_opts(
                title_opts=opts.TitleOpts(title='现金流入占比分析'),
                xaxis_opts=opts.AxisOpts(name='年份', name_location='middle', name_gap=25),
                yaxis_opts=opts.AxisOpts(name='现 \n金 \n流 \n入 \n占 \n比 \n(%)', name_location='middle', name_gap=40,
                                         name_rotate=0, axislabel_opts=opts.LabelOpts(formatter='{value}%')),
                legend_opts=opts.LegendOpts(type_='scroll', pos_left='25%', pos_right='20%'),
                datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                tooltip_opts=opts.TooltipOpts(trigger='axis',
                                              formatter=formatter_bar,
                                              axis_pointer_type='none'),
                toolbox_opts=opts.ToolboxOpts(
                    feature=opts.ToolBoxFeatureOpts(
                        data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                        brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                        magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['line', 'stack'])
                    )
                )
            )

        )

        line = (
            Line()
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='经营活动占比同比增长率',
                           y_axis=yoy1[- (len(yoy1) - 1):],
                           yaxis_index=1, is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[0]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='投资活动占比同比增长率',
                           y_axis=yoy2[- (len(yoy2) - 1):],
                           yaxis_index=1,
                           is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[1]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='筹资活动占比同比增长率',
                           y_axis=yoy3[- (len(yoy3) - 1):],
                           yaxis_index=1, is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[2]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=formatter_line_label
                                                           ),
                                 linestyle_opts=opts.LineStyleOpts(width=WIDTH)
                                 )
        )

        bar.overlap(line)

        return bar

    def get_outflow(self) -> Bar:
        """
        :return: 现金流出分析图(Bar_Overlap_Line)
        """
        df = DataProcess(self.ticker).cash_flow_statement_data()

        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '%' + '<br />';
            }
            return form;
            }
            """
        )

        formatter_line = JsCode(
            """function(params) {
            form =
            params.name + '年' + '<br />'
            + '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:'
            + params.color + '"></span>'
            + params.seriesName + '：'
            + params.data[1] + '%';

            return form;
        }"""
        )

        formatter_line_label = JsCode(
            """function(params) {
            form = params.value[1] + '%';
            return form;
        }"""
        )

        x = df.index[- (len(df.index) - 1):]
        data1 = df['经营活动现金流出小计']
        data2 = df['投资活动现金流出小计']
        data3 = df['筹资活动现金流出小计']
        tot = data1 + data2 + data3
        d1 = (data1 / tot * 100).round(2)
        d2 = (data2 / tot * 100).round(2)
        d3 = (data3 / tot * 100).round(2)
        yoy1 = list(((d1 - d1.shift(1)) / d1.shift(1) * 100).round(2))  # 占比同比增长率
        yoy2 = list(((d2 - d2.shift(1)) / d2.shift(1) * 100).round(2))
        yoy3 = list(((d3 - d3.shift(1)) / d3.shift(1) * 100).round(2))

        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='outflow'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='经营活动现金流出占比',
                           y_axis=list(d1[- (len(d1) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.8),
                           gap=0)
                .add_yaxis(series_name='投资活动现金流出占比',
                           y_axis=list(d2[- (len(d2) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2], opacity=0.8),
                           gap=0)
                .add_yaxis(series_name='筹资活动现金流出占比',
                           y_axis=list(d3[- (len(d3) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], opacity=0.8),
                           gap=0)
                .extend_axis(
                yaxis=opts.AxisOpts(name=' 现\n 金\n 流\n 出\n 同\n 比\n 增\n 长\n 率\n(%)', name_location='middle', name_gap=50,
                                    name_rotate=0, axislabel_opts=opts.LabelOpts(formatter='{value}%')))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=JsCode(
                                                               "function(params){return Number(params.data).toFixed(1) + '%';}"
                                                           ))
                                 )
                .set_global_opts(
                title_opts=opts.TitleOpts(title='现金流出占比分析'),
                xaxis_opts=opts.AxisOpts(name='年份', name_location='middle', name_gap=25),
                yaxis_opts=opts.AxisOpts(name='现 \n金 \n流 \n出 \n占 \n比 \n(%)', name_location='middle', name_gap=40,
                                         name_rotate=0, axislabel_opts=opts.LabelOpts(formatter='{value}%')),
                legend_opts=opts.LegendOpts(type_='scroll', pos_left='25%', pos_right='20%'),
                datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                tooltip_opts=opts.TooltipOpts(trigger='axis',
                                              formatter=formatter_bar,
                                              axis_pointer_type='none'),
                toolbox_opts=opts.ToolboxOpts(
                    feature=opts.ToolBoxFeatureOpts(
                        data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                        brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                        magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['line', 'stack'])
                    )
                )
            )

        )

        line = (
            Line()
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='经营活动占比同比增长率',
                           y_axis=yoy1[- (len(yoy1) - 1):],
                           yaxis_index=1, is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[0]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='投资活动占比同比增长率',
                           y_axis=yoy2[- (len(yoy2) - 1):],
                           yaxis_index=1, is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[1]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='筹资活动占比同比增长率',
                           y_axis=yoy3[- (len(yoy3) - 1):],
                           yaxis_index=1, is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[2]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=formatter_line_label
                                                           ),
                                 linestyle_opts=opts.LineStyleOpts(width=WIDTH)
                                 )
        )

        bar.overlap(line)

        return bar

    # 之前弄得麻烦了，可以sankey表是用timeline组合起来的，对timeline设置chart_id才对
    # @staticmethod
    # def paste_id() -> NoReturn:
    #     """
    #     寻找html中sankey表的id（因不是标准echarts图表，所以无法指定id）
    #     :return: NoReturn
    #     """
    #     soup = BeautifulSoup(open(r'.\data\raw_figure_html\现金流量表分析_raw.html', encoding='utf-8'), features="lxml")
    #     div_lst = soup.find_all('div', class_='chart-container')
    #     id_lst = []
    #
    #     # 获得html的id列表
    #     for i in div_lst:
    #         id = re.search("(?<=id=\").+?(?=\")", str(i)).group()
    #         id_lst.append(id)
    #     # 寻找sankey图的id
    #     for id in id_lst:
    #         if len(id) > 10:
    #             sankey_id = id
    #         else:
    #             pass
    #     # 替换sankey图的id
    #     # 读
    #     with open(r'.\data\config_data\chart_config_CashFlowStatement.json', 'rb') as r:
    #         params = json.load(r)
    #         for dic in params:
    #             if len(dic['cid']) > 10:
    #                 dic['cid'] = sankey_id
    #             else:
    #                 pass
    #     r.close()
    #     # 写
    #     with open(r'.\data\config_data\chart_config_CashFlowStatement.json', 'w') as w:
    #         json.dump(params, w)
    #     w.close()

    def plot(self) -> NoReturn:
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title='现金流量表', layout=Page.DraggablePageLayout)
        page.add(self.get_netflow()[0], self.get_netflow()[1], self.get_netflow()[2], self.get_netflow()[3],
                 self.get_sankey(), self.get_inflow(), self.get_outflow())
        raw_path = r'.\data\raw_figure_html\现金流量表分析_raw.html'
        cfg_path = r'.\data\config_data\chart_config_CashFlowStatement.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_现金流量表分析.html'
        page.render(raw_path)
        # self.paste_id()
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class StatementOfProfitAndLoss:
    def __init__(self, ticker):
        self.ticker = ticker

    def get_MB_profit(self) -> Bar:
        """
        :return: 主营业务利润分析图
        """
        df = DataProcess(self.ticker).statement_of_profit_and_loss_data()
        # 设置文本格式,java回调函数
        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '亿元' + '<br />';
            }
            return form;
            }
            """
        )
        formatter_line = JsCode(
            """function(params) {
            form =
            params.name + '年' + '<br />'
            + '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:'
            + params.color 
            + '"></span>'
            + params.seriesName + '：'
            + params.data[1] + '%';

            return form;
        }"""
        )
        formatter_line_label = JsCode(
            """function(params) {
            form = params.value[1] + '%';
            return form;
        }"""
        )

        x = df.index[- (len(df.index) - 1):]
        d1 = df['主营营业收入'] / 1e8
        d2 = df['主营营业支出'] / 1e8
        d3 = (df['主营营业收入'] - df['主营营业支出']) / 1e8
        yoy1 = list(((d1 - d1.shift(1)) / d1.shift(1) * 100).round(2))  # 占比同比增长率
        yoy2 = list(((d2 - d2.shift(1)) / d2.shift(1) * 100).round(2))
        yoy3 = list(((d3 - d3.shift(1)) / d3.shift(1) * 100).round(2))
        d1 = d1.round(3)
        d2 = d2.round(3)
        d3 = d3.round(3)

        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='MB_profit'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='主营营业收入',
                           y_axis=list(d1[- (len(d1) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.75),
                           gap=0)
                .add_yaxis(series_name='主营营业支出',
                           y_axis=list(d2[- (len(d2) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], opacity=0.75),
                           gap=0)
                .add_yaxis(series_name='主营营业利润',
                           y_axis=list(d3[- (len(d3) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2], opacity=0.75),
                           gap=0)
                .extend_axis(
                yaxis=opts.AxisOpts(name=' 同\n 比\n 增\n 长\n 率',
                                    name_location='middle',
                                    name_gap=50,
                                    name_rotate=0,
                                    axislabel_opts=opts.LabelOpts(formatter='{value}%')))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=JsCode(
                                                               "function(params){return Number(params.data).toFixed(1) + '亿';}"
                                                           ))
                                 )
                .set_global_opts(title_opts=opts.TitleOpts(title='主营业务利润分析', subtitle='单位: 亿元'),
                                 xaxis_opts=opts.AxisOpts(name='年份',
                                                          name_location='middle',
                                                          name_gap=25),
                                 yaxis_opts=opts.AxisOpts(name='数 \n额 ',
                                                          name_location='middle',
                                                          name_gap=40,
                                                          name_rotate=0),
                                 legend_opts=opts.LegendOpts(type_='scroll',
                                                             pos_left='20%',
                                                             pos_right='20%'),
                                 datazoom_opts=[opts.DataZoomOpts(range_start=0,
                                                                  range_end=100,
                                                                  pos_bottom='-7px')],
                                 tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                               formatter=formatter_bar,
                                                               axis_pointer_type='none'),
                                 toolbox_opts=opts.ToolboxOpts(
                                     feature=opts.ToolBoxFeatureOpts(
                                         data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                                         brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                                         magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['line']))
                                 )
                                 )
        )

        line = (
            Line()
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='主营营业收入同比增长率',
                           y_axis=yoy1[- (len(yoy1) - 1):],
                           yaxis_index=1,
                           is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[0]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='主营营业支出同比增长率',
                           y_axis=yoy2[- (len(yoy2) - 1):],
                           yaxis_index=1,
                           is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[1]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='主营营业利润同比增长率',
                           y_axis=yoy3[- (len(yoy3) - 1):],
                           yaxis_index=1,
                           is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[2]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=formatter_line_label
                                                           ),
                                 linestyle_opts=opts.LineStyleOpts(width=WIDTH)
                                 )
        )

        bar.overlap(line)
        return bar

    def get_profit(self) -> Bar:
        """
        :return: 利润分析图
        """
        df = DataProcess(self.ticker).statement_of_profit_and_loss_data()

        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '亿元' + '<br />';
            }
            return form;
            }
            """
        )
        formatter_line = JsCode(
            """function(params) {
            form =
            params.name + '年' + '<br />'
            + '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:'
            + params.color 
            + '"></span>'
            + params.seriesName + '：'
            + params.data[1] + '%';

            return form;
        }"""
        )
        formatter_line_label = JsCode(
            """function(params) {
            form = params.value[1] + '%';
            return form;
        }"""
        )

        x = df.index[- (len(df.index) - 1):]
        d1 = df['营业收入'] / 1e8
        d2 = df['营业成本'] / 1e8
        d3 = df['净利润'] / 1e8
        yoy1 = list(((d1 - d1.shift(1)) / d1.shift(1) * 100).round(2))  # 占比同比增长率
        yoy2 = list(((d2 - d2.shift(1)) / d2.shift(1) * 100).round(2))
        yoy3 = list(((d3 - d3.shift(1)) / d3.shift(1) * 100).round(2))
        d1 = d1.round(3)
        d2 = d2.round(3)
        d3 = d3.round(3)

        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='profit'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='营业收入',
                           y_axis=list(d1[- (len(d1) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.75),
                           gap=0)
                .add_yaxis(series_name='营业成本',
                           y_axis=list(d2[- (len(d2) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], opacity=0.75),
                           gap=0)
                .add_yaxis(series_name='净利润',
                           y_axis=list(d3[- (len(d3) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2], opacity=0.75),
                           gap=0)
                .extend_axis(
                yaxis=opts.AxisOpts(name=' 同\n 比\n 增\n 长\n 率',
                                    name_location='middle',
                                    name_gap=50,
                                    name_rotate=0,
                                    axislabel_opts=opts.LabelOpts(formatter='{value}%')))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=JsCode(
                                                               "function(params){return Number(params.data).toFixed(1) + '亿';}"
                                                           ))
                                 )
                .set_global_opts(title_opts=opts.TitleOpts(title='利润分析', subtitle='单位: 亿元'),
                                 xaxis_opts=opts.AxisOpts(name='年份',
                                                          name_location='middle',
                                                          name_gap=25),
                                 yaxis_opts=opts.AxisOpts(name='数 \n额 ',
                                                          name_location='middle',
                                                          name_gap=40,
                                                          name_rotate=0),
                                 legend_opts=opts.LegendOpts(type_='scroll',
                                                             pos_left='20%',
                                                             pos_right='20%'),
                                 datazoom_opts=[opts.DataZoomOpts(range_start=0,
                                                                  range_end=100,
                                                                  pos_bottom='-7px')],
                                 tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                               formatter=formatter_bar,
                                                               axis_pointer_type='none'),
                                 toolbox_opts=opts.ToolboxOpts(
                                     feature=opts.ToolBoxFeatureOpts(
                                         data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                                         brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                                         magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['line']))
                                 )
                                 )
        )

        line = (
            Line()
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='营业收入同比增长率',
                           y_axis=yoy1[- (len(yoy1) - 1):],
                           yaxis_index=1,
                           is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[0]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='营业成本同比增长率',
                           y_axis=yoy2[- (len(yoy2) - 1):],
                           yaxis_index=1,
                           is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[1]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .add_yaxis(series_name='净利润同比增长率',
                           y_axis=yoy3[- (len(yoy3) - 1):],
                           yaxis_index=1,
                           is_selected=True,
                           itemstyle_opts=opts.ItemStyleOpts(color=BACK_LINE_COLOR_LST[2]),
                           tooltip_opts=opts.TooltipOpts(formatter=formatter_line))
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='top',
                                                           font_size=10,
                                                           formatter=formatter_line_label
                                                           ),
                                 linestyle_opts=opts.LineStyleOpts(width=WIDTH)
                                 )
        )

        bar.overlap(line)
        return bar

    def get_cost(self) -> Bar:
        """
        :return: 费用分析图
        """
        df = DataProcess(self.ticker).statement_of_profit_and_loss_data()

        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '亿元' + '<br />';
            }
            return form;
            }
            """
        )

        x_bar = df.index[- (len(df.index) - 1):]
        d1 = (df['研发费用'] / 1e8).round(3)
        d2 = (df['销售费用'] / 1e8).round(3)
        d3 = (df['管理费用'] / 1e8).round(3)
        d4 = (df['财务费用'] / 1e8).round(3)

        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='cost'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(x_bar)])
                .add_yaxis(series_name='研发费用',
                           y_axis=list(d1[- (len(d1) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2], opacity=0.8),
                           gap=0)
                .add_yaxis(series_name='销售费用',
                           y_axis=list(d2[- (len(d2) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.8),
                           gap=0)
                .add_yaxis(series_name='管理费用',
                           y_axis=list(d3[- (len(d3) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], opacity=0.8),
                           gap=0)
                .add_yaxis(series_name='财务费用',
                           y_axis=list(d4[- (len(d4) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[3], opacity=0.8),
                           gap=0)
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           font_size=10,
                                                           formatter=JsCode(
                                                               "function(params){return Number(params.data).toFixed(1) + '亿';}"
                                                           ))
                                 )
                .set_global_opts(title_opts=opts.TitleOpts(title='费用分析', subtitle='单位: 亿元'),
                                 xaxis_opts=opts.AxisOpts(name='年份',
                                                          name_location='end',
                                                          name_gap=25,
                                                          type_='category'),
                                 yaxis_opts=opts.AxisOpts(name='数  \n额  ',
                                                          name_location='middle',
                                                          name_gap=40,
                                                          name_rotate=0,
                                                          splitline_opts=opts.SplitLineOpts(is_show=True)),
                                 datazoom_opts=[opts.DataZoomOpts(range_start=0,
                                                                  range_end=100,
                                                                  pos_bottom='-7px')],
                                 tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                               formatter=formatter_bar,
                                                               axis_pointer_type='none'),
                                 toolbox_opts=opts.ToolboxOpts(
                                     feature=opts.ToolBoxFeatureOpts(
                                         data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                                         brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                                         magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['line', 'stack']))
                                 )
                                 )
        )
        return bar

    def get_income(self) -> Bar:
        """
        收入分析图(主营业务 与 其他业务)
        :return: Bar_with_Stack
        """
        df = DataProcess(self.ticker).statement_of_profit_and_loss_data()

        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '亿元' + '<br />';
            }
            return form;
            }
            """
        )

        x = df.index[- (len(df.index) - 1):]
        d1 = (df['主营营业收入'] / 1e8).round(3)
        d2 = (df['其他业务收入(附注)'] / 1e8).round(3)
        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='income'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='主营营业收入',
                           y_axis=list(d1[- (len(d1) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.9),
                           gap=0,
                           category_gap='50%',
                           stack='stack1')
                .add_yaxis(series_name='其他业务收入',
                           y_axis=list(d2[- (len(d2) - 1):]),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[3], opacity=0.9),
                           gap=0,
                           category_gap='50%',
                           stack='stack1')
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='right',
                                                           font_size=13,
                                                           formatter=JsCode(
                                                               "function(params){return Number(params.data).toFixed(1) + '亿';}"
                                                           ))
                                 )
                .set_global_opts(title_opts=opts.TitleOpts(title='收入分析', subtitle='单位: 亿元'),
                                 xaxis_opts=opts.AxisOpts(name='年份',
                                                          name_location='middle',
                                                          name_gap=25,
                                                          type_='category'),
                                 yaxis_opts=opts.AxisOpts(name='数  \n额  ',
                                                          name_location='middle',
                                                          name_gap=40,
                                                          name_rotate=0,
                                                          splitline_opts=opts.SplitLineOpts(is_show=True)),
                                 datazoom_opts=[opts.DataZoomOpts(range_start=0,
                                                                  range_end=100,
                                                                  pos_bottom='-7px')],
                                 tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                               formatter=formatter_bar,
                                                               axis_pointer_type='none')
                                 )
        )
        return bar

    def plot(self) -> NoReturn:
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="利润表", layout=Page.DraggablePageLayout)
        page.add(self.get_profit(), self.get_MB_profit(), self.get_income(), self.get_cost())
        raw_path = r'.\data\raw_figure_html\利润表分析_raw.html'
        cfg_path = r'.\data\config_data\chart_config_StatementOfProfitAndLoss.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_利润表分析.html'
        page.render(raw_path)
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class StatementOfFinancialPosition:
    def __init__(self, ticker):
        self.ticker = ticker

    def get_abs(self) -> Timeline:
        """
        :return: 资产负债绝对值时间线图
        """
        df = DataProcess(self.ticker).statement_of_financial_position()
        tl = Timeline(init_opts=opts.InitOpts(chart_id='abs'))
        for date in df.index:
            d1 = float(df.loc[date, '资产总计']) / 1e8
            d2 = float(df.loc[date, '流动资产合计']) / 1e8
            d3 = float(df.loc[date, '非流动资产合计']) / 1e8
            if d1 != 0:
                prop2 = float(df.loc[date, '流动资产合计'] / df.loc[date, '资产总计'])  # 流动资产占比
                prop3 = float(df.loc[date, '非流动资产合计'] / df.loc[date, '资产总计'])  # 非流动资产占比
            else:
                prop2 = 0
                prop3 = 0

            d4 = float(df.loc[date, '负债合计']) / 1e8
            d5 = float(df.loc[date, '流动负债合计']) / 1e8
            d6 = float(df.loc[date, '非流动负债合计']) / 1e8
            # d7 = float(df.loc[date, '股东权益合计']) / 1e8
            if d4 != 0:
                prop5 = float(df.loc[date, '流动负债合计'] / df.loc[date, '负债合计'])  # 流动负债占比
                prop6 = float(df.loc[date, '非流动负债合计'] / df.loc[date, '负债合计'])  # 非流动负债占比
            else:
                prop5 = 0
                prop6 = 0

            fontsize = 18

            formatter_liquid1 = JsCode("function(x){return x.seriesName + '\\n' + '" + str(
                round(d1, 1)) + "' + '亿' + '\\n' + '\\n' + '100%';}")
            lq1 = (
                Liquid()
                    .add('资产总计', [1], color=[COLOR_LST[1]], center=["16.6%", "25%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid1, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            formatter_liquid2 = JsCode(
                "function(x){return x.seriesName + '\\n' + '" + str(round(d2, 1)) + "' + '亿' + '\\n' + '\\n' + '" + str(
                    int(prop2 * 100)) + "' + '%';}")
            lq2 = (
                Liquid()
                    .add('流动资产合计', [prop2], color=[COLOR_LST[1]], center=["50%", "25%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid2, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            formatter_liquid3 = JsCode(
                "function(x){return x.seriesName + '\\n' + '" + str(round(d3, 1)) + "' + '亿' + '\\n' + '\\n' + '" + str(
                    int(prop3 * 100)) + "' + '%';}")
            lq3 = (
                Liquid()
                    .add('非流动资产合计', [prop3], color=[COLOR_LST[1]], center=["83.3%", "25%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid3, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            formatter_liquid4 = JsCode("function(x){return x.seriesName + '\\n' + '" + str(
                round(d4, 1)) + "' + '亿' + '\\n' + '\\n' + '100%';}")
            lq4 = (
                Liquid()
                    .add('负债合计', [1], color=[COLOR_LST[1]], center=["16.6%", "75%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid4, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            formatter_liquid5 = JsCode(
                "function(x){return x.seriesName + '\\n' + '" + str(round(d5, 1)) + "' + '亿' + '\\n' + '\\n' + '" + str(
                    int(prop5 * 100)) + "' + '%';}")
            lq5 = (
                Liquid()
                    .add('流动负债合计', [prop5], color=[COLOR_LST[1]], center=["50%", "75%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid5, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            formatter_liquid6 = JsCode(
                "function(x){return x.seriesName + '\\n' + '" + str(round(d6, 1)) + "' + '亿' + '\\n' + '\\n' + '" + str(
                    int(prop6 * 100)) + "' + '%';}")
            lq6 = (
                Liquid()
                    .add('非流动负债合计', [prop6], color=[COLOR_LST[1]], center=["83.3%", "75%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid6, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            grid = (
                Grid()
                    .add(lq1, grid_opts=opts.GridOpts())
                    .add(lq2, grid_opts=opts.GridOpts())
                    .add(lq3, grid_opts=opts.GridOpts())
                    .add(lq4, grid_opts=opts.GridOpts())
                    .add(lq5, grid_opts=opts.GridOpts())
                    .add(lq6, grid_opts=opts.GridOpts())
            )
            tl.add(grid, f"{date[:4]}")

        tl_width = 40
        tl.add(Pie(), '').add_schema(play_interval=1000,
                                     is_loop_play=False,
                                     pos_bottom='-8px',
                                     pos_left=str(tl_width * 0.7) + '%',
                                     width=str(tl_width) + '%',
                                     controlstyle_opts=opts.TimelineControlStyle(item_size=18))

        return tl

    def get_equity_liability(self) -> Bar:
        """
        :return: 负债和权益的堆叠柱状图
        """
        df = DataProcess(self.ticker).statement_of_financial_position()

        formatter_bar = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data + '亿元' + '<br />';
            }
            return form;
            }
            """
        )

        x = df.index
        d1 = (df['负债合计'] / 1e8).round(3)
        d2 = (df['股东权益合计'] / 1e8).round(3)
        bar = (
            Bar(init_opts=opts.InitOpts(theme=THEME, chart_id='equity_liability'))
                .add_xaxis(xaxis_data=[i[:4] for i in list(x)])
                .add_yaxis(series_name='负债合计',
                           y_axis=list(d1),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1], opacity=0.9),
                           gap=0,
                           category_gap='70%',
                           stack='stack1')
                .add_yaxis(series_name='股东权益合计',
                           y_axis=list(d2),
                           itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[0], opacity=0.9),
                           gap=0,
                           category_gap='70%',
                           stack='stack1')
                .set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                           position='right',
                                                           font_size=13,
                                                           formatter=JsCode(
                                                               "function(params){return Number(params.data).toFixed(1) + '亿';}"
                                                           ))
                                 )
                .set_global_opts(title_opts=opts.TitleOpts(title='负债股权比例分析', subtitle='单位: 亿元'),
                                 xaxis_opts=opts.AxisOpts(name='年份',
                                                          name_location='middle',
                                                          name_gap=25,
                                                          type_='category'),
                                 yaxis_opts=opts.AxisOpts(name='数\n额',
                                                          name_location='middle',
                                                          name_gap=50,
                                                          name_rotate=0,
                                                          splitline_opts=opts.SplitLineOpts(is_show=True)),
                                 datazoom_opts=[opts.DataZoomOpts(range_start=0,
                                                                  range_end=100,
                                                                  pos_bottom='-7px')],
                                 tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                               formatter=formatter_bar,
                                                               axis_pointer_type='none')
                                 )
        )
        return bar

    def get_equity(self) -> Timeline:
        """
        :return: 权益细节的比例饼状时间序列图
        """
        df = DataProcess(self.ticker).statement_of_financial_position()

        formatter_pie = JsCode(
            """function(params){
                color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                return params.seriesName + '<br />' + color_icon + params.color + '"></span>' + params.name + ': ' + Number(params.value).toFixed(3) + '亿';
             }
            """
        )

        tl = Timeline(init_opts=opts.InitOpts(chart_id='equity'))

        x1 = ['流动资产合计', '非流动资产合计']
        x2 = ['货币资金', '应收票据及应收账款', '预付款项', '存货', '其他流动资产']
        x3 = ['长期股权投资', '固定资产', '在建工程', '无形资产', '其他非流动资产']
        for date in df.index:
            y_lst = []
            for i in [x1, x2, x3]:
                y = []
                for j in range(len(i)):
                    y.append(float(df.loc[date, i[j]]) / 1e8)
                y_lst.append(y)
            y1, y2, y3 = y_lst[0], y_lst[1], y_lst[2]

            data_pair_lst = []
            for i, j in [(x1, y1), (x2, y2), (x3, y3)]:
                data_pair = [list(z) for z in zip(i, j)]
                if i == x1:
                    pass
                else:
                    data_pair.sort(key=lambda x: x[1])
                data_pair_lst.append(data_pair)
            data_pair1, data_pair2, data_pair3 = data_pair_lst[0], data_pair_lst[1], data_pair_lst[2]

            pie = (
                Pie()
                    .add(series_name='总资产比例',
                         data_pair=data_pair1,
                         radius='45%',  # 30
                         center=['50%', '50%'],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .add(series_name='流动资产比例',
                         data_pair=data_pair2,
                         radius=['37%', '47%'],  # 22, 32
                         center=['22%', '50%'],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .add(series_name='非流动资产比例',
                         data_pair=data_pair3,
                         radius=['37%', '47%'],  # 22, 32
                         center=['78%', '50%'],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .set_colors(PIE_COLOR_LST)
                    .set_series_opts(label_opts=opts.LabelOpts(font_size=10, formatter="{b}\n{d}%"))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title='资产比例分析'),
                    legend_opts=opts.LegendOpts(type_='scroll', pos_left='25%', pos_right='20%', pos_top='10%')
                )
            )

            tl.add(pie, f"{date[:4]}")

        tl.add_schema(play_interval=1000, is_loop_play=False, pos_bottom='10%')
        return tl

    def get_liability(self) -> Timeline:
        """
        :return: 负债细节的比例饼状时间序列图
        """
        df = DataProcess(self.ticker).statement_of_financial_position()

        formatter_pie = JsCode(
            """function(params){
                color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                return params.seriesName + '<br />' + color_icon + params.color + '"></span>' + params.name + ': ' + Number(params.value).toFixed(3) + '亿';
             }
            """
        )

        tl = Timeline(init_opts=opts.InitOpts(chart_id='liability'))

        x1 = ['流动负债合计', '非流动负债合计']
        x2 = ['短期借款', '应付票据及应付账款', '合同负债', '应付职工薪酬', '其他流动负债']
        x3 = ['长期借款', '应付债券', '租赁负债', '递延收益', '其他非流动负债']
        for date in df.index:
            y_lst = []
            for i in [x1, x2, x3]:
                y = []
                for j in range(len(i)):
                    y.append(float(df.loc[date, i[j]]) / 1e8)
                y_lst.append(y)
            y1, y2, y3 = y_lst[0], y_lst[1], y_lst[2]

            data_pair_lst = []
            for i, j in [(x1, y1), (x2, y2), (x3, y3)]:
                data_pair = [list(z) for z in zip(i, j)]
                if i == x1:
                    pass
                else:
                    data_pair.sort(key=lambda x: x[1])
                data_pair_lst.append(data_pair)
            data_pair1, data_pair2, data_pair3 = data_pair_lst[0], data_pair_lst[1], data_pair_lst[2]

            pie = (
                Pie()
                    .add(series_name='总负债比例',
                         data_pair=data_pair1,
                         radius='45%',
                         center=['50%', '50%'],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .add(series_name='流动负债比例',
                         data_pair=data_pair2,
                         radius=['37%', '47%'],
                         center=['22%', '50%'],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .add(series_name='非流动负债比例',
                         data_pair=data_pair3,
                         radius=['37%', '47%'],
                         center=['78%', '50%'],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .set_colors(PIE_COLOR_LST)
                    .set_series_opts(label_opts=opts.LabelOpts(font_size=10, formatter="{b}\n{d}%"))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title='负债比例分析'),
                    legend_opts=opts.LegendOpts(type_='scroll', pos_left='25%', pos_right='20%', pos_top='10%')
                )
            )

            tl.add(pie, f"{date[:4]}")
        tl.add_schema(play_interval=1000, is_loop_play=False, pos_bottom='10%')
        return tl

    def plot(self):
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="资产负债表", layout=Page.DraggablePageLayout)
        page.add(self.get_abs(), self.get_equity_liability(), self.get_equity(), self.get_liability())
        raw_path = r'.\data\raw_figure_html\资产负债表分析_raw.html'
        cfg_path = r'.\data\config_data\chart_config_StatementOfFinancialPosition.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_资产负债表分析.html'
        page.render(raw_path)
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class Solvency:  # 数值位数有问题
    def __init__(self, ticker):
        self.ticker = ticker

    def all_figures(self) -> dict:
        """
        偿债能力分析的六张图
        :return: 含[Line, Table]数据的字典
        """
        df = DataProcess(self.ticker).solvency()[0]
        df_ind = DataProcess(self.ticker).solvency()[1]
        rename_dic = dict(zip(df_ind.columns, df.columns))
        df_ind.rename(columns=rename_dic, inplace=True)
        x = [i[:4] for i in list(df.index)]

        c_dic = {}
        count = 0

        formatter_line_label = JsCode(
            """function(params) {
            form = params.value[1] + '%';
            return form;
        }"""
        )

        row_dic = {}  # 总表的字典

        for name in df.columns.drop('CODES'):
            # 折线图数据
            y = [round(i, 2) for i in list(df[name])]
            y_ind = [round(i, 2) for i in list(df_ind[name])]

            # 总表的设置
            overall_headers = ['差额百分比'] + [i[:4] for i in list(df.index)]
            row_dic[name] = [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                             [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

            # 表格数据
            if name == '资产负债率':
                headers = [''] + [i[:4] for i in list(df.index)]
                row1 = ['企业'] + [str(i) + '%' for i in y]
                row2 = ['行业'] + [str(i) + '%' for i in y_ind]
                row3 = ['差额'] + [str(k) + '%' for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                row4 = ['差额百分比'] + [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                                    [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '%' + '<br />';
                    }
                    return form;
                    }
                    """
                )
            else:
                headers = [''] + [i[:4] for i in list(df.index)]
                row1 = ['企业'] + [str(i) for i in y]
                row2 = ['行业'] + [str(i) for i in y_ind]
                row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                row4 = ['差额百分比'] + [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                                    [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '<br />';
                    }
                    return form;
                    }
                    """
                )

            line = (
                Line(init_opts=opts.InitOpts(chart_id='id_' + str(count)))
                    .add_xaxis(xaxis_data=x)
                    .add_yaxis(series_name='企业',
                               y_axis=y,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='行业',
                               y_axis=y_ind,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='差额(企业-行业)',
                               y_axis=[round(y[i] - y_ind[i], 2) for i in range(len(y))],
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[3]),
                               is_connect_nones=True,
                               label_opts=opts.LabelOpts(color=COLOR_LST[3])
                               )
                    # .set_series_opts(areastyle_opts=opts.AreaStyleOpts(opacity=0.65))
                    .set_series_opts(linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title=name),
                    datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                    xaxis_opts=opts.AxisOpts(name='年份',
                                             name_location='middle',
                                             name_gap=25,
                                             boundary_gap=False),
                    yaxis_opts=opts.AxisOpts(name='数\n值',
                                             name_location='middle',
                                             name_gap=40,
                                             name_rotate=0,
                                             splitline_opts=opts.SplitLineOpts(is_show=True)),
                    tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                  formatter=formatter_line,
                                                  axis_pointer_type='cross'),
                    legend_opts=opts.LegendOpts(pos_top='6.5%', pos_left='20%'),
                    toolbox_opts=opts.ToolboxOpts(
                        feature=opts.ToolBoxFeatureOpts(
                            data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                            brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                            magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                        )
                    )
                )
            )

            table = (
                Table().add(headers=headers, rows=[row1, row2, row3, row4])
            )

            if name == '资产负债率':
                line.set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                               position='top',
                                                               font_size=12,
                                                               formatter=formatter_line_label
                                                               )
                                     )

            c_dic[name] = [line, table]
            count += 1

        # 总表
        overall_table = Table()
        overall_rows = []
        for key, val in row_dic.items():
            overall_rows.append([key] + val)
        overall_table.add(headers=overall_headers, rows=overall_rows)

        return [overall_table, c_dic]

    @staticmethod
    def paste_id() -> NoReturn:
        """
        寻找html中Table的id（因不是标准echarts图表，所以无法指定id）
        :return: NoReturn
        """
        soup = BeautifulSoup(open(r'.\data\raw_figure_html\偿债能力分析_raw.html', encoding='utf-8'), features="lxml")
        div_lst = soup.find_all('div', class_='chart-container')
        id_lst = []

        # 获得html的id列表
        for i in div_lst:
            id = re.search("(?<=id=\").+?(?=\")", str(i)).group()
            id_lst.append(id)
        # 寻找Table的id
        table_id_lst = []
        for id in id_lst:
            if len(id) > 10:
                table_id_lst.append(id)
            else:
                pass
        # 替换Table的id
        # 读
        with open(r'.\data\config_data\chart_config_Solvency.json', 'rb') as r:
            count = 0
            params = json.load(r)
            for dic in params:
                if len(dic['cid']) > 10:
                    dic['cid'] = table_id_lst[count]
                    count += 1
                else:
                    pass
        r.close()
        # 写
        with open(r'.\data\config_data\chart_config_Solvency.json', 'w') as w:
            json.dump(params, w)
        w.close()

    def plot(self):
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="偿债能力", layout=Page.DraggablePageLayout)
        df = DataProcess(self.ticker).solvency()[0]
        for name in df.columns.drop('CODES'):
            page.add(self.all_figures()[1][name][0]).add(self.all_figures()[1][name][1])
        page.add(self.all_figures()[0])  # 画总表
        raw_path = r'.\data\raw_figure_html\偿债能力分析_raw.html'
        cfg_path = r'.\data\config_data\chart_config_Solvency.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_偿债能力分析.html'
        page.render(raw_path)
        Solvency(self.ticker).paste_id()
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class GrowthAbility:
    def __init__(self, ticker):
        self.ticker = ticker

    def all_figures(self):
        """
        成长能力分析的六张图
        :return: 含[Line, Table]数据的字典
        """
        df = DataProcess(self.ticker).growth_ability()[0]
        df_ind = DataProcess(self.ticker).growth_ability()[1]
        rename_dic = dict(zip(df_ind.columns, df.columns))
        df_ind.rename(columns=rename_dic, inplace=True)
        x = [i[:4] for i in list(df.index)]

        c_dic = {}
        row_dic = {}

        formatter_line = JsCode(
            """function(params){
            color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
            form = params[0].name + '年' + '<br />';
            for(i=0; i<params.length; i++){
            form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '%' + '<br />';
            }
            return form;
            }
            """
        )
        formatter_line_label = JsCode(
            """function(params) {
            form = params.value[1] + '%';
            return form;
        }"""
        )

        for name in df.columns.drop('CODES'):
            # 折线图数据
            y = [round(i, 2) for i in list(df[name])]
            y_ind = [round(i, 2) for i in list(df_ind[name])]

            # 总表的设置
            overall_headers = ['差额百分比'] + [i[:4] for i in list(df.index)]
            row_dic[name] = [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                             [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

            # 表格数据
            headers = [''] + [i[:4] for i in list(df.index)]
            row1 = ['企业'] + [str(i) + '%' for i in y]
            row2 = ['行业'] + [str(i) + '%' for i in y_ind]
            row3 = ['差额'] + [str(k) + '%' for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
            row4 = ['差额百分比'] + [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                                [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

            line = (
                Line(init_opts=opts.InitOpts(chart_id=name))
                    .add_xaxis(xaxis_data=x)
                    .add_yaxis(series_name='企业',
                               y_axis=y,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='行业',
                               y_axis=y_ind,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='差额(企业-行业)',
                               y_axis=[round(y[i] - y_ind[i], 2) for i in range(len(y))],
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[3]),
                               is_connect_nones=True)
                    # .set_series_opts(areastyle_opts=opts.AreaStyleOpts(opacity=0.65),
                    #                  label_opts=opts.LabelOpts(is_show=False))
                    .set_series_opts(label_opts=opts.LabelOpts(formatter=formatter_line_label),
                                     linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title=name),
                    datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                    xaxis_opts=opts.AxisOpts(name='年份',
                                             name_location='middle',
                                             name_gap=25,
                                             boundary_gap=False),
                    yaxis_opts=opts.AxisOpts(name='数\n值',
                                             name_location='middle',
                                             name_gap=40,
                                             name_rotate=0,
                                             axislabel_opts=opts.LabelOpts(formatter="{value}%"),
                                             splitline_opts=opts.SplitLineOpts(is_show=True)),
                    tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                  formatter=formatter_line,
                                                  axis_pointer_type='cross'),
                    legend_opts=opts.LegendOpts(pos_top='6.5%', pos_left='20%'),
                    toolbox_opts=opts.ToolboxOpts(
                        feature=opts.ToolBoxFeatureOpts(
                            data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                            brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                            magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                        )
                    )
                )
            )

            table = (
                Table().add(headers=headers,
                            rows=[row1, row2, row3, row4])
            )

            c_dic[name] = [line, table]

        # 总表
        overall_table = Table()
        overall_rows = []
        for key, val in row_dic.items():
            overall_rows.append([key] + val)
        overall_table.add(headers=overall_headers, rows=overall_rows)

        return [overall_table, c_dic]

    @staticmethod
    def paste_id() -> NoReturn:
        """
        寻找html中Table的id（因不是标准echarts图表，所以无法指定id）
        :return: NoReturn
        """
        soup = BeautifulSoup(open(r'.\data\raw_figure_html\成长能力分析_raw.html', encoding='utf-8'), features="lxml")
        div_lst = soup.find_all('div', class_='chart-container')
        id_lst = []

        # 获得html的id列表
        for i in div_lst:
            id = re.search("(?<=id=\").+?(?=\")", str(i)).group()
            id_lst.append(id)
        # 寻找Table的id
        table_id_lst = []
        for id in id_lst:
            if len(id) > 10:
                table_id_lst.append(id)
            else:
                pass
        # 替换Table的id
        # 读
        with open(r'.\data\config_data\chart_config_GrowthAbility.json', 'rb') as r:
            count = 0
            params = json.load(r)
            for dic in params:
                if len(dic['cid']) > 10:
                    dic['cid'] = table_id_lst[count]
                    count += 1
                else:
                    pass
        r.close()
        # 写
        with open(r'.\data\config_data\chart_config_GrowthAbility.json', 'w') as w:
            json.dump(params, w)
        w.close()

    def plot(self):
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="成长能力", layout=Page.DraggablePageLayout)
        df = DataProcess(self.ticker).growth_ability()[0]
        for name in df.columns.drop('CODES'):
            page.add(self.all_figures()[1][name][0]).add(self.all_figures()[1][name][1])
        page.add(self.all_figures()[0])  # 画总表
        raw_path = r'.\data\raw_figure_html\成长能力分析_raw.html'
        cfg_path = r'.\data\config_data\chart_config_GrowthAbility.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_成长能力分析.html'
        page.render(raw_path)
        GrowthAbility(self.ticker).paste_id()
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class Profitability:
    def __init__(self, ticker):
        self.ticker = ticker

    def all_figures(self):
        """
        盈利能力分析的六张图
        :return: 含[Line, Table]数据的字典
        """
        df = DataProcess(self.ticker).profitability()[0]
        df_ind = DataProcess(self.ticker).profitability()[1]
        rename_dic = dict(zip(df_ind.columns, df.columns))
        df_ind.rename(columns=rename_dic, inplace=True)
        x = [i[:4] for i in list(df.index)]

        c_dic = {}
        row_dic = {}

        for name in df.columns:
            # 折线图数据
            y = [round(i, 2) for i in list(df[name])]
            y_ind = [round(i, 2) for i in list(df_ind[name])]

            # 总表的设置
            overall_headers = ['差额百分比'] + [i[:4] for i in list(df.index)]
            row_dic[name] = [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                             [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

            # 表格数据
            if name == '每股收益EPS' or name == '每股营业总收入':
                headers = ['单位:元'] + [i[:4] for i in list(df.index)]
                row1 = ['企业'] + [str(i) for i in y]
                row2 = ['行业'] + [str(i) for i in y_ind]
                row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                row4 = ['差额百分比'] + [
                    str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                    [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                subtitle = '单位: 元'
                formatter_yaxis = '{value}'

                formatter_line_label = JsCode(
                    """function(params) {
                    form = params.value[1] + '元';
                    return form;
                }"""
                )
                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '元' + '<br />';
                    }
                    return form;
                    }
                    """
                )

            else:
                headers = [''] + [i[:4] for i in list(df.index)]
                row1 = ['企业'] + [str(i) + '%' for i in y]
                row2 = ['行业'] + [str(i) + '%' for i in y_ind]
                row3 = ['差额'] + [str(k) + '%' for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                row4 = ['差额百分比'] + [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                                    [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                subtitle = ''
                formatter_yaxis = '{value}%'

                formatter_line_label = JsCode(
                    """function(params) {
                    form = params.value[1] + '%';
                    return form;
                }"""
                )
                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '%' + '<br />';
                    }
                    return form;
                    }
                    """
                )

            line = (
                Line(init_opts=opts.InitOpts(chart_id=name))
                    .add_xaxis(xaxis_data=x)
                    .add_yaxis(series_name='企业',
                               y_axis=y,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='行业',
                               y_axis=y_ind,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='差额(企业-行业)',
                               y_axis=[round(y[i] - y_ind[i], 2) for i in range(len(y))],
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[3]),
                               is_connect_nones=True)
                    # .set_series_opts(areastyle_opts=opts.AreaStyleOpts(opacity=0.65),
                    #                  label_opts=opts.LabelOpts(is_show=False))
                    .set_series_opts(label_opts=opts.LabelOpts(formatter=formatter_line_label),
                                     linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title=name, subtitle=subtitle),
                    datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                    xaxis_opts=opts.AxisOpts(name='年份',
                                             name_location='middle',
                                             name_gap=25,
                                             boundary_gap=False),
                    yaxis_opts=opts.AxisOpts(name='数\n值',
                                             name_location='middle',
                                             name_gap=40,
                                             name_rotate=0,
                                             axislabel_opts=opts.LabelOpts(formatter=formatter_yaxis),
                                             splitline_opts=opts.SplitLineOpts(is_show=True)),
                    tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                  formatter=formatter_line,
                                                  axis_pointer_type='cross'),
                    legend_opts=opts.LegendOpts(pos_top='6.5%', pos_left='20%'),
                    toolbox_opts=opts.ToolboxOpts(
                        feature=opts.ToolBoxFeatureOpts(
                            data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                            brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                            magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                        )
                    )
                )
            )

            table = (
                Table().add(headers=headers,
                            rows=[row1, row2, row3, row4])
            )

            c_dic[name] = [line, table]

        # 总表
        overall_table = Table()
        overall_rows = []
        for key, val in row_dic.items():
            overall_rows.append([key] + val)
        overall_table.add(headers=overall_headers, rows=overall_rows)

        return [overall_table, c_dic]

    @staticmethod
    def paste_id() -> NoReturn:
        """
        寻找html中Table的id（因不是标准echarts图表，所以无法指定id）
        :return: NoReturn
        """
        soup = BeautifulSoup(open(r'.\data\raw_figure_html\盈利能力分析_raw.html', encoding='utf-8'), features="lxml")
        div_lst = soup.find_all('div', class_='chart-container')
        id_lst = []

        # 获得html的id列表
        for i in div_lst:
            id = re.search("(?<=id=\").+?(?=\")", str(i)).group()
            id_lst.append(id)
        # 寻找Table的id
        table_id_lst = []
        for id in id_lst:
            if len(id) > 10:
                table_id_lst.append(id)
            else:
                pass
        # 替换Table的id
        # 读
        with open(r'.\data\config_data\chart_config_Profitability.json', 'rb') as r:
            count = 0
            params = json.load(r)
            for dic in params:
                if len(dic['cid']) > 10:
                    dic['cid'] = table_id_lst[count]
                    count += 1
                else:
                    pass
        r.close()
        # 写
        with open(r'.\data\config_data\chart_config_Profitability.json', 'w') as w:
            json.dump(params, w)
        w.close()

    def plot(self):
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="盈利能力", layout=Page.DraggablePageLayout)
        df = DataProcess(self.ticker).profitability()[0]
        for name in df.columns:
            page.add(self.all_figures()[1][name][0]).add(self.all_figures()[1][name][1])
        page.add(self.all_figures()[0])  # 画总表
        raw_path = r'.\data\raw_figure_html\盈利能力分析_raw.html'
        cfg_path = r'.\data\config_data\chart_config_Profitability.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_盈利能力分析.html'
        page.render(raw_path)
        Profitability(self.ticker).paste_id()
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class OperatingCapacity:
    def __init__(self, ticker):
        self.ticker = ticker

    def get_abs(self) -> Timeline:
        """
        绝对值（存货、应收票据及应收账款、流动资产合计）水滴图
        :return:水滴图Timeline
        """

        df = DataProcess(self.ticker).statement_of_financial_position()
        tl = Timeline(init_opts=opts.InitOpts(chart_id='abs'))

        for date in df.index:
            d3 = float(df.loc[date, '流动资产合计']) / 1e8
            prop3 = float(df.loc[date, '流动资产合计']) / float(df.loc[:, '流动资产合计'].max())
            d1 = float(df.loc[date, '存货']) / 1e8
            prop1 = float(df.loc[date, '存货'] / df.loc[date, '流动资产合计']) * prop3
            d2 = float(df.loc[date, '应收票据及应收账款']) / 1e8
            prop2 = float(df.loc[date, '应收票据及应收账款'] / df.loc[date, '流动资产合计']) * prop3

            fontsize = 18

            formatter_liquid1 = JsCode(
                "function(x){return x.seriesName + '\\n' + '" + str(round(d1, 1)) + "' + '亿';}")
            lq1 = (
                Liquid()
                    .add('存货', [prop1], color=[COLOR_LST[1]], center=["83.3%", "50%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid1, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            formatter_liquid2 = JsCode(
                "function(x){return x.seriesName + '\\n' + '" + str(round(d2, 1)) + "' + '亿';}")
            lq2 = (
                Liquid()
                    .add('应收票据及应收账款', [prop2], color=[COLOR_LST[1]], center=["50%", "50%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid2, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            formatter_liquid3 = JsCode(
                "function(x){return x.seriesName + '\\n' + '" + str(round(d3, 1)) + "' + '亿';}")
            lq3 = (
                Liquid()
                    .add('流动资产合计', [prop3], color=[COLOR_LST[1]], center=["16.6%", "50%"], is_outline_show=False,
                         label_opts=opts.LabelOpts(font_size=fontsize, formatter=formatter_liquid3, position='inside'))
                    .set_global_opts(tooltip_opts=opts.TooltipOpts(is_show=False))
            )

            grid = (
                Grid()
                    .add(lq1, grid_opts=opts.GridOpts())
                    .add(lq2, grid_opts=opts.GridOpts())
                    .add(lq3, grid_opts=opts.GridOpts())
            )
            tl.add(grid, f"{date[:4]}")

        tl.add(Pie(), '').add_schema(play_interval=1000, is_loop_play=False, pos_bottom='7%')

        return tl

    def all_figures(self) -> dict:
        """
        营运能力分析的六张图
        :return: 含[Line, Table]数据的字典
        """
        df = DataProcess(self.ticker).operating_capacity()[0]
        df_ind = DataProcess(self.ticker).operating_capacity()[1]
        rename_dic = dict(zip(df_ind.columns, df.columns))
        df_ind.rename(columns=rename_dic, inplace=True)
        x = [i[:4] for i in list(df.index)]

        c_dic = {}
        row_dic = {}

        formatter_line_label = JsCode(
            """function(params) {
            form = params.value[1] + '天';
            return form;
        }"""
        )

        for name in df.columns.drop('CODES'):
            # 折线图数据
            y = [round(i, 2) for i in list(df[name])]
            y_ind = [round(i, 2) for i in list(df_ind[name])]

            # 总表的设置
            overall_headers = ['差额百分比'] + [i[:4] for i in list(df.index)]
            row_dic[name] = [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                             [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

            # 表格数据
            if name == '营业周期':
                headers = ['单位:天'] + [i[:4] for i in list(df.index)]
                row1 = ['企业'] + [str(i) for i in y]
                row2 = ['行业'] + [str(i) for i in y_ind]
                row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                row4 = ['差额百分比'] + [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                                    [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                # label_formatter = "{value}天"
                subtitle = '单位: 天'

                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '天' + '<br />';
                    }
                    return form;
                    }
                    """
                )

            else:
                headers = [''] + [i[:4] for i in list(df.index)]
                row1 = ['企业'] + [str(i) for i in y]
                row2 = ['行业'] + [str(i) for i in y_ind]
                row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                row4 = ['差额百分比'] + [str(round(k * 100, 2)) + '%' if abs(k * 100) < ANOMALOUS_BOUNDARY else '*' + str(round(k * 100, 2)) + '%' for k in
                                    [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                # label_formatter = "{value}"
                subtitle = '单位: 次'

                formatter_line = JsCode(
                    """function(params){
                    color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                    form = params[0].name + '年' + '<br />';
                    for(i=0; i<params.length; i++){
                    form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '次' + '<br />';
                    }
                    return form;
                    }
                    """
                )

            line = (
                Line(init_opts=opts.InitOpts(chart_id=name))
                    .add_xaxis(xaxis_data=x)
                    .add_yaxis(series_name='企业',
                               y_axis=y,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[1]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='行业',
                               y_axis=y_ind,
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[2]),
                               is_connect_nones=True)
                    .add_yaxis(series_name='差额(企业-行业)',
                               y_axis=[round(y[i] - y_ind[i], 2) for i in range(len(y))],
                               itemstyle_opts=opts.ItemStyleOpts(color=COLOR_LST[3]),
                               is_connect_nones=True,
                               label_opts=opts.LabelOpts(color=COLOR_LST[3]))
                    # .set_series_opts(
                    # areastyle_opts=opts.AreaStyleOpts(opacity=0.65),
                    # label_opts=opts.LabelOpts(is_show=False))
                    .set_series_opts(linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title=name, subtitle=subtitle),
                    datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                    xaxis_opts=opts.AxisOpts(name='年份',
                                             name_location='middle',
                                             name_gap=25,
                                             boundary_gap=False),
                    yaxis_opts=opts.AxisOpts(name='数\n值',
                                             name_location='middle',
                                             name_gap=40,
                                             name_rotate=0,
                                             # axislabel_opts=opts.LabelOpts(formatter=label_formatter),
                                             splitline_opts=opts.SplitLineOpts(is_show=True)),
                    tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                  formatter=formatter_line,
                                                  axis_pointer_type='cross'),
                    legend_opts=opts.LegendOpts(pos_top='6.5%', pos_left='20%'),
                    toolbox_opts=opts.ToolboxOpts(
                        feature=opts.ToolBoxFeatureOpts(
                            data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                            brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                            magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                        )
                    )
                )
            )

            if name == '营业周期':
                line.set_series_opts(label_opts=opts.LabelOpts(formatter=formatter_line_label))

            table = (
                Table().add(headers=headers,
                            rows=[row1, row2, row3, row4])
            )

            c_dic[name] = [line, table]

        # 总表
        overall_table = Table()
        overall_rows = []
        for key, val in row_dic.items():
            overall_rows.append([key] + val)
        overall_table.add(headers=overall_headers, rows=overall_rows)

        return [overall_table, c_dic]

    def get_prop(self):
        df1 = DataProcess(self.ticker).operating_capacity()[2]
        df2 = DataProcess(self.ticker).operating_capacity()[3]

        x1 = df1.columns
        x2 = df2.columns

        tl1 = Timeline()
        tl2 = Timeline()

        formatter_pie = JsCode(
            """function(params){
                color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                return params.seriesName + '<br />' + color_icon + params.color + '"></span>' + params.name + ': ' + Number(params.value).toFixed(3) + '亿';
             }
            """
        )

        for date in df1.index:
            y1 = (df1.loc[date, :] / 1e8).round(6)
            data_pair1 = [list(z) for z in zip(x1, y1)]

            pie1 = (
                Pie(init_opts=opts.InitOpts(chart_id=str(date) + 'pie1'))
                    .add("应收账款明细", data_pair=data_pair1,
                         center=['50%', '50%'],
                         radius=[90, 105],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .set_colors(PIE_COLOR_LST[7:])
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title='应收账款明细', subtitle='单位:亿元'),
                    legend_opts=opts.LegendOpts(type_='scroll', pos_top='20%', pos_left='80%', orient='vertical')
                )
            )
            tl1.add(pie1, f"{date[:4]}年").add_schema(pos_bottom='7%')

        for date in df2.index:
            y2 = (df2.loc[date, :] / 1e8).round(6)
            data_pair2 = [list(z) for z in zip(x2, y2)]

            pie2 = (
                Pie(init_opts=opts.InitOpts(chart_id=str(date) + 'pie2'))
                    .add("存货项目明细",
                         data_pair=data_pair2,
                         center=['50%', '50%'],
                         radius=[90, 105],
                         tooltip_opts=opts.TooltipOpts(formatter=formatter_pie))
                    .set_colors(PIE_COLOR_LST[2:])
                    .set_global_opts(
                    title_opts=opts.TitleOpts(title='存货项目明细', subtitle='单位:亿元'),
                    legend_opts=opts.LegendOpts(type_='scroll', pos_top='20%', pos_left='80%', orient='vertical')
                )
            )
            tl2.add(pie2, f"{date[:4]}年").add_schema(pos_bottom='7%')

        return [tl1, tl2]

    @staticmethod
    def paste_id() -> NoReturn:
        """
        寻找html中Table的id（因不是标准echarts图表，所以无法指定id）
        :return: NoReturn
        """
        soup = BeautifulSoup(open(r'.\data\raw_figure_html\营运能力分析_raw.html', encoding='utf-8'), features="lxml")
        div_lst = soup.find_all('div', class_='chart-container')
        id_lst = []

        # 获得html的id列表
        for i in div_lst:
            id = re.search("(?<=id=\").+?(?=\")", str(i)).group()
            id_lst.append(id)
        # 寻找Table的id
        table_id_lst = []
        for id in id_lst:
            if len(id) > 10:
                table_id_lst.append(id)
            else:
                pass
        # 替换Table的id
        # 读
        with open(r'.\data\config_data\chart_config_OperatingCapacity.json', 'rb') as r:
            count = 0
            params = json.load(r)
            for dic in params:
                if len(dic['cid']) > 10:
                    dic['cid'] = table_id_lst[count]
                    count += 1
                else:
                    pass
        r.close()
        # 写
        with open(r'.\data\config_data\chart_config_OperatingCapacity.json', 'w') as w:
            json.dump(params, w)
        w.close()

    def plot(self):
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="营运能力", layout=Page.DraggablePageLayout)
        df = DataProcess(self.ticker).operating_capacity()[0]
        page.add(self.get_abs())
        for name in df.columns.drop('CODES'):
            page.add(self.all_figures()[1][name][0]).add(self.all_figures()[1][name][1])
        page.add(self.all_figures()[0])  # 画总表
        page.add(self.get_prop()[0]).add(self.get_prop()[1])
        raw_path = r'.\data\raw_figure_html\营运能力分析_raw.html'
        cfg_path = r'.\data\config_data\chart_config_OperatingCapacity.json'
        ripe_path = r'.\data\figure_html\\' + self.ticker + '_营运能力分析.html'
        page.render(raw_path)
        OperatingCapacity(self.ticker).paste_id()
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


class ExcelAnalysis1:
    def __init__(self, ticker):
        self.ticker = ticker

    def write_calculation_table(self):
        # 企业数据
        df1 = DataProcess(self.ticker).excel_analysis1()[0]  # 读取数据
        df1.drop(labels=['CODES'], axis=1, inplace=True)  # 删除第一列（股票代码）
        df1.index = [z[: 4] for z in df1.index]  # 时间标签年化（2021-12-31 -> 2021）
        df1.rename(mapper=lambda x: str(x) + '_企业', axis=0, inplace=True)  # 年份标签重命名（2021 -> 2021_企业）

        # 行业数据
        df2 = DataProcess(self.ticker).excel_analysis1()[1]
        df2.drop(labels=['CODES'], axis=1, inplace=True)
        df2.index = [z[: 4] for z in df2.index]
        df2.rename(mapper=lambda x: str(x) + '_行业', axis=0, inplace=True)

        path = r'./data/excel_template/excel_eco_template.xlsx'
        wb = load_workbook(path)

        # 企业数据写入模板 #
        ws1 = wb[wb.sheetnames[0]]

        col_warehouse1 = [z for z in string.ascii_uppercase]
        col_warehouse1.remove('A')
        # 年份修改
        for i, col in enumerate(col_warehouse1[: len(df1.index)]):
            coordinate = col + '1'
            ws1[coordinate].value = df1.index.tolist()[i]

        data_rows_index = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 19, 20, 21, 23, 24, 26, 27, 28, 29, 30, 31]  # 需要写入数据的行
        data_rows_name = ['应收账款增长率', '应收票据增长率', '其他应收款合计增长率', '预付款项增长率', '存货增长率',
                          '在建工程增长率', '长期待摊费用增长率', '固定资产增长率', '商誉增长率', '资产减值准备增长率',
                          '短期借款增长率', '应付票据增长率', '应付账款增长率', '其他应付款合计增长率', '营业总收入增长率',
                          '主营营业收入增长率', '营业总成本增长率', '销售毛利率增长率', '销售净利率增长率', '存货增长率',
                          '固定资产增长率', '应收账款比率增长率', '存货比率增长率', '固定资产比率增长率', '经营活动产生的现金流量净额增长率']
        if len(data_rows_index) != len(data_rows_name):
            logger.error('ExcelAnalysis1-企业-数据写入位置数与指标数不对应！')
        else:
            data_rows = dict(zip(data_rows_index, data_rows_name))
            for row, name in data_rows.items():
                for i, col in enumerate(col_warehouse1[: len(df1.index)]):
                    coordinate = col + str(row)
                    cell = ws1[coordinate]
                    cell.value = df1[name][i]

        # 企业vs行业数据写入模板 #
        ws2 = wb[wb.sheetnames[2]]

        col_warehouse21 = [z for z in string.ascii_uppercase]
        col_warehouse21.remove('A')
        col_warehouse21 = col_warehouse21[: : 2]
        col_warehouse22 = [z for z in string.ascii_uppercase]
        col_warehouse22.remove('A')
        col_warehouse22 = col_warehouse22[1: : 2]
        # 年份修改
        for i, col in enumerate(col_warehouse21[: len(df1.index)]):
            coordinate = col + '1'
            ws2[coordinate].value = df1.index.tolist()[i]
        for j, col in enumerate(col_warehouse22[: len(df2.index)]):
            coordinate = col + '1'
            ws2[coordinate].value = df2.index.tolist()[j]

        data_rows_index_2 = [3, 4, 5, 6, 8, 10, 11, 13, 14, 15, 17, 18, 19, 20, 21, 22, 24, 25]
        data_rows_name_2 = ['预付款项增长率', '应收账款增长率', '存货增长率', '固定资产增长率', '应付账款增长率', '总资产周转率',
                            '净资产收益率', '销售毛利率', '销售净利率', '费用率', '应收账款周转率', '存货周转率',
                            '固定资产周转率', '应收账款比率增长率', '存货比率增长率', '固定资产比率增长率', '货币资金/总资产', '资产负债率']
        if len(data_rows_index_2) != len(data_rows_name_2):
            logger.error('ExcelAnalysis1-企业vs行业-数据写入位置数与指标数不对应！')
        else:
            data_rows_2 = dict(zip(data_rows_index_2, data_rows_name_2))
            for row, name in data_rows_2.items():
                for i, col in enumerate(col_warehouse21[: len(df1.index)]):
                    coordinate = col + str(row)
                    cell = ws2[coordinate]
                    cell.value = df1[name][i]
                for j, col in enumerate(col_warehouse22[: len(df2.index)]):
                    coordinate = col + str(row)
                    cell = ws2[coordinate]
                    cell.value = df2[name][j]

        save_path = r'.\data\excel_data\\' + self.ticker + '_舞弊财务指标异常风险分析.xlsx'
        wb.save(save_path)

    def write_exhibition_table(self):
        # 判断后的数据
        df = DataProcess(self.ticker).excel_analysis1_judge_law()  # 读取数据
        df.drop(labels=['CODES'], axis=1, inplace=True)  # 删除第一列（股票代码）
        df.index = [z[: 4] for z in df.index]  # 时间标签年化（2021-12-31 -> 2021）

        base_path = r'.\data\excel_data\\' + self.ticker + '_舞弊财务指标异常风险分析.xlsx'
        wb = load_workbook(base_path)

        # 企业自身比较判断结果写入
        ws1 = wb[wb.sheetnames[1]]

        col_warehouse = [z for z in string.ascii_uppercase]
        col_warehouse.remove('A')
        # 年份修改
        for i, col in enumerate(col_warehouse[: len(df.index)]):
            coordinate = col + '1'
            ws1[coordinate].value = df.index.tolist()[i]

        style_g = NamedStyle(name='style_g', fill=PatternFill(patternType='solid', fgColor='6C85F3'))
        style_z = NamedStyle(name='style_z', fill=PatternFill(patternType='solid', fgColor='50B8DF'))
        style_d = NamedStyle(name='style_d', fill=PatternFill(patternType='solid', fgColor='AED7F2'))

        data_rows_index = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 19, 20, 21, 23, 24, 26, 27, 28, 29, 30, 31]  # 需要写入数据的行
        data_rows_name = ['应收账款增长率', '应收票据增长率', '其他应收款合计增长率', '预付款项增长率', '存货增长率',
                          '在建工程增长率', '长期待摊费用增长率', '固定资产增长率', '商誉增长率', '资产减值准备增长率',
                          '短期借款增长率', '应付票据增长率', '应付账款增长率', '其他应付款合计增长率', '营业总收入增长率',
                          '主营营业收入增长率', '营业总成本增长率', '销售毛利率增长率', '销售净利率增长率', '存货增长率',
                          '固定资产增长率', '应收账款比率增长率', '存货比率增长率', '固定资产比率增长率', '经营活动产生的现金流量净额增长率']
        data_rows_name = ['qy' + z for z in data_rows_name]
        data_rows = dict(zip(data_rows_index, data_rows_name))
        for row, name in data_rows.items():
            for i, col in enumerate(col_warehouse[: len(df.index)]):
                coordinate = col + str(row)
                cell = ws1[coordinate]
                cell.value = df[name][i]
                if cell.value == '高':
                    cell.style = style_g
                elif cell.value == '中':
                    cell.style = style_z
                elif cell.value == '低':
                    cell.style = style_d

        # 企业vs行业数据写入
        ws2 = wb[wb.sheetnames[3]]

        # 年份修改
        for i, col in enumerate(col_warehouse[: len(df.index)]):
            coordinate = col + '1'
            ws2[coordinate].value = df.index.tolist()[i]

        data_rows_index_2 = [3, 4, 5, 6, 8, 10, 11, 13, 14, 15, 17, 18, 19, 20, 21, 22, 24, 25]
        data_rows_name_2 = ['预付款项增长率', '应收账款增长率', '存货增长率', '固定资产增长率', '应付账款增长率', '总资产周转率',
                            '净资产收益率', '销售毛利率', '销售净利率', '费用率', '应收账款周转率', '存货周转率',
                            '固定资产周转率', '应收账款比率增长率', '存货比率增长率', '固定资产比率增长率', '货币资金/总资产', '资产负债率']
        data_rows_name_2 = ['vs' + z for z in data_rows_name_2]
        data_rows_2 = dict(zip(data_rows_index_2, data_rows_name_2))
        for row, name in data_rows_2.items():
            for i, col in enumerate(col_warehouse[: len(df.index)]):
                coordinate = col + str(row)
                cell = ws2[coordinate]
                cell.value = df[name][i]
                if cell.value == '高':
                    cell.style = style_g
                elif cell.value == '中':
                    cell.style = style_z
                elif cell.value == '低':
                    cell.style = style_d

        save_path = r'.\data\excel_data\\' + self.ticker + '_舞弊财务指标异常风险分析.xlsx'
        wb.save(save_path)

    def write_all(self):
        self.write_calculation_table()
        self.write_exhibition_table()
        file_path = r'.\data\excel_data\\' + self.ticker + '_舞弊财务指标异常风险分析.xlsx'  # 等于save_path
        os.startfile(file_path)


# excel表格分析部分
class ExcelAnalysis2:
    def __init__(self, ticker):
        self.ticker = ticker

    def write_calculation_table(self):
        df = DataProcess(self.ticker).excel_analysis2()  # 读取数据
        df = df.iloc[4:, :]  # 前4年数据是多的，DataProcess返回数据因为要做判断，所以没有删掉前四年数据
        df.drop(labels=['CODES'], axis=1, inplace=True)  # 删除第一列（股票代码）
        df.index = [z[: 4] for z in df.index]  # 时间标签年化（2021-12-31 -> 2021）

        template_path = r'./data/excel_template/excel_template.xlsx'
        wb = load_workbook(template_path)
        ws = wb[wb.sheetnames[0]]


        col_warehouse = [z for z in string.ascii_uppercase]
        col_warehouse.remove('A')
        # 年份修改
        for i, col in enumerate(col_warehouse[: len(df.index)]):
            coordinate = col + '1'
            ws[coordinate].value = df.index.tolist()[i]

        data_rows_index = [4, 7, 8, 12, 15, 16, 17, 20, 23, 24, 25, 29, 30, 32, 34, 35, 37,
                           38, 40, 42, 43, 45, 46, 48, 49, 51, 52, 53, 54, 56, 57, 58, 59]  # 需要写入数据的行
        data_rows_name = ['固定资产占总资产的比重变化率', '固定资产增加率', '累计折旧占固定资产原值的比重变化率', '存货占总资产比重变化率',
                          '存货增加率', '营业收入增长率', '营业成本增长率', '存货跌价准备占存货的比重变化率', '应收账款增长率', '存货增长率',
                          '营业收入增长率', '现金及现金等价物占总资产的比重变化率', '应收账款占总资产的比重增长率', '坏账占应收账款比重变化率',
                          '销售费用增长速度', '营业收入增长速度', '研发费用增长率', '营业收入增长率', '毛利率变化率', '毛利率变化率', '存货周转率变化率',
                          '毛利率变化率', '应收账款周转率变化率', '毛利率变化率', '应付账款变化率', '销售毛利率', '现金循环周期', '行业毛利率',
                          '行业现金循环周期', '销售毛利率', '销售费用率', '行业毛利率', '行业销售费用率']
        data_rows = dict(zip(data_rows_index, data_rows_name))
        for row, name in data_rows.items():
            for i, col in enumerate(col_warehouse[: len(df.index)]):
                coordinate = col + str(row)
                cell = ws[coordinate]
                cell.value = df[name][i]

        print(1)
        save_path = r'.\data\excel_data\\' + self.ticker + '_舞弊勾稽指标异常风险分析.xlsx'
        wb.save(save_path)

    def write_exhibition_table(self):
        df = DataProcess(self.ticker).excel_analysis2_judge_law()  # 读取数据
        df.drop(labels=['CODES'], axis=1, inplace=True)  # 删除第一列（股票代码）
        df.index = [z[: 4] for z in df.index]  # 时间标签年化（2021-12-31 -> 2021）

        base_path = r'.\data\excel_data\\' + self.ticker + '_舞弊勾稽指标异常风险分析.xlsx'
        wb = load_workbook(base_path)
        ws = wb[wb.sheetnames[1]]

        col_warehouse_new = [z for z in string.ascii_uppercase]
        # 年份修改
        for i, col in enumerate(col_warehouse_new[: len(df.index)]):
            coordinate = col + '1'
            ws[coordinate].value = df.index.tolist()[i]

        style_g = NamedStyle(name='style_g', fill=PatternFill(patternType='solid', fgColor='6C85F3'))
        style_z = NamedStyle(name='style_z', fill=PatternFill(patternType='solid', fgColor='50B8DF'))
        style_d = NamedStyle(name='style_d', fill=PatternFill(patternType='solid', fgColor='AED7F2'))

        data_rows_index = [4, 6, 9, 11, 13, 15, 17, 19, 22, 24, 26, 28, 30, 32, 34, 36, 38, 40]  # 需要写入数据的行
        data_rows_name = ['law11', 'law12', 'law21', 'law22', 'law23', 'law24', 'law25', 'law26', 'law31',
                          'law32', 'law33', 'law34', 'law35', 'law36', 'law37', 'law38', 'law39', 'law310']
        data_rows = dict(zip(data_rows_index, data_rows_name))
        for row, name in data_rows.items():
            for i, col in enumerate(col_warehouse_new[: len(df.index)]):
                coordinate = col + str(row)
                cell = ws[coordinate]
                cell.value = df[name][i]
                if cell.value == '高':
                    cell.style = style_g
                elif cell.value == '中':
                    cell.style = style_z
                elif cell.value == '低':
                    cell.style = style_d

        save_path = r'.\data\excel_data\\' + self.ticker + '_舞弊勾稽指标异常风险分析.xlsx'
        wb.save(save_path)

    def write_all(self):
        self.write_calculation_table()
        self.write_exhibition_table()
        file_path = r'.\data\excel_data\\' + self.ticker + '_舞弊勾稽指标异常风险分析.xlsx'  # 等于save_path
        os.startfile(file_path)


class PlotAll():
    def __init__(self, ticker):
        self.ticker = ticker

    def plot(self):
        try:
            BasicInfo(self.ticker).plot()
            StatementOfFinancialPosition(self.ticker).plot()
            CashFlowStatement(self.ticker).plot()
            StatementOfProfitAndLoss(self.ticker).plot()
            Solvency(self.ticker).plot()
            OperatingCapacity(self.ticker).plot()
            Profitability(self.ticker).plot()
            GrowthAbility(self.ticker).plot()
            logger.info('绘图成功！')
            ExcelAnalysis1(self.ticker).write_all()
            ExcelAnalysis2(self.ticker).write_all()
            logger.info('excel分析表输出成功！')
        except BaseException as e:
            logger.error(e)


class ExtraPlot():
    def __init__(self, ticker):
        self.ticker = ticker

    def trial_figures(self):
        if self.ticker == '601857.SH':
            # part1
            df_o = DataProcess(self.ticker).operating_capacity()[0]
            df_o_ind = DataProcess(self.ticker).operating_capacity()[1]
            rename_dic = dict(zip(df_o_ind.columns, df_o.columns))
            df_o_ind.rename(columns=rename_dic, inplace=True)
            x = [i[:4] for i in list(df_o.index)]

            c_dic = {}
            temp_color_lst = ['', '#FFC0CB', '#00FFFF', '#EEEEEE']
            level_val = {1: '优秀值', 2: '良好值', 3: '平均值', 4: '较低值', 5: '较差值'}
            level_color = ['', '#7C4DFF', '#306AFE', '#2979FF', '#2196F3', '#64B5F6']

            for name in ['应收账款周转率', '总资产周转率']:  # '存货周转率'
                # 折线图数据
                y = [round(i, 2) for i in list(df_o[name])]
                y_ind = [round(i, 2) for i in list(df_o_ind[name])]

                # 表格数据
                if name == '营业周期':
                    headers = ['单位:天'] + [i[:4] for i in list(df_o.index)]
                    row1 = ['企业'] + [str(i) for i in y]
                    row2 = ['行业'] + [str(i) for i in y_ind]
                    row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                    row4 = ['差额百分比'] + [
                        str(round(k * 100, 2)) + '%' if abs(k * 100) < 50 else '*' + str(round(k * 100, 2)) + '%' for k
                        in
                        [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                    # label_formatter = "{value}天"
                    subtitle = '单位: 天'

                    formatter_line = JsCode(
                        """function(params){
                        color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                        form = params[0].name + '年' + '<br />';
                        for(i=0; i<params.length; i++){
                        form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '天' + '<br />';
                        }
                        return form;
                        }
                        """
                    )

                else:
                    headers = [''] + [i[:4] for i in list(df_o.index)]
                    row1 = ['企业'] + [str(i) for i in y]
                    row2 = ['行业'] + [str(i) for i in y_ind]
                    row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                    row4 = ['差额百分比'] + [
                        str(round(k * 100, 2)) + '%' if abs(k * 100) < 50 else '*' + str(round(k * 100, 2)) + '%' for k
                        in
                        [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                    # label_formatter = "{value}"
                    subtitle = '单位: 次'

                    formatter_line = JsCode(
                        """function(params){
                        color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                        form = params[0].name + '年' + '<br />';
                        for(i=0; i<params.length; i++){
                        form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '次' + '<br />';
                        }
                        return form;
                        }
                        """
                    )

                line = (
                    Line(init_opts=opts.InitOpts(chart_id=name))
                        .add_xaxis(xaxis_data=x)
                        .add_yaxis(series_name='企业',
                                   y_axis=y,
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[1]),
                                   is_connect_nones=True)
                        .add_yaxis(series_name='行业',
                                   y_axis=y_ind,
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[2]),
                                   is_connect_nones=True)
                        .add_yaxis(series_name='差额(企业-行业)',
                                   y_axis=[round(y[i] - y_ind[i], 2) for i in range(len(y))],
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[3]),
                                   is_connect_nones=True,
                                   label_opts=opts.LabelOpts(color=temp_color_lst[3]))
                        # .set_series_opts(
                        # areastyle_opts=opts.AreaStyleOpts(opacity=0.65),
                        # label_opts=opts.LabelOpts(is_show=False))
                        .set_series_opts(linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                        .set_global_opts(
                        title_opts=opts.TitleOpts(title=name, subtitle=subtitle),
                        datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                        xaxis_opts=opts.AxisOpts(name='年份',
                                                 name_location='middle',
                                                 name_gap=25,
                                                 boundary_gap=False),
                        yaxis_opts=opts.AxisOpts(name='数\n值',
                                                 name_location='middle',
                                                 name_gap=40,
                                                 name_rotate=0,
                                                 # axislabel_opts=opts.LabelOpts(formatter=label_formatter),
                                                 splitline_opts=opts.SplitLineOpts(is_show=True)),
                        tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                      formatter=formatter_line,
                                                      axis_pointer_type='cross'),
                        legend_opts=opts.LegendOpts(type_='scroll', pos_left='20%', pos_right='20%'),
                        toolbox_opts=opts.ToolboxOpts(
                            feature=opts.ToolBoxFeatureOpts(
                                data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                                brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                                magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                            )
                        )
                    )
                )

                table = (
                    Table().add(headers=headers,
                                rows=[row1, row2, row3, row4])
                )

                if name.find('应收账款周转率') >= 0:
                    dic_extra_o = DataProcess(self.ticker).get_extra_trial_data()
                    level_line = Line().add_xaxis(x)
                    for level in range(1, 6):
                        level_line.add_yaxis(level_val[level],
                                             dic_extra_o['ReceivablesTurnover'][str(level)],
                                             is_symbol_show=False,
                                             label_opts=opts.LabelOpts(is_show=False),
                                             areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
                                             itemstyle_opts=opts.ItemStyleOpts(color=level_color[level])
                                             )
                    line.overlap(level_line)
                elif name.find('存货周转率') >= 0:
                    dic_extra_o = DataProcess(self.ticker).get_extra_trial_data()
                    level_line = Line().add_xaxis(x)
                    for level in range(1, 6):
                        level_line.add_yaxis(level_val[level],
                                             dic_extra_o['InventoryTurnover'][str(level)],
                                             is_symbol_show=False,
                                             label_opts=opts.LabelOpts(is_show=False),
                                             areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
                                             itemstyle_opts=opts.ItemStyleOpts(color=level_color[level])
                                             )
                    line.overlap(level_line)
                elif name.find('总资产周转率') >= 0:
                    dic_extra_o = DataProcess(self.ticker).get_extra_trial_data()
                    level_line = Line().add_xaxis(x)
                    for level in range(1, 6):
                        level_line.add_yaxis(level_val[level],
                                             dic_extra_o['AssetTurnover'][str(level)],
                                             is_symbol_show=False,
                                             label_opts=opts.LabelOpts(is_show=False),
                                             areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
                                             itemstyle_opts=opts.ItemStyleOpts(color=level_color[level])
                                             )
                    line.overlap(level_line)

                c_dic[name] = [line, table]

            # part2
            df_s = DataProcess(self.ticker).solvency()[0]
            df_s_ind = DataProcess(self.ticker).solvency()[1]
            rename_dic = dict(zip(df_s_ind.columns, df_s.columns))
            df_s_ind.rename(columns=rename_dic, inplace=True)
            x = [i[:4] for i in list(df_s.index)]

            count = 0

            formatter_line_label = JsCode(
                """function(params) {
                form = params.value[1] + '%';
                return form;
            }"""
            )

            for name in ['资产负债率', '已获利息倍数']:
                # 折线图数据
                y = [round(i, 2) for i in list(df_s[name])]
                y_ind = [round(i, 2) for i in list(df_s_ind[name])]

                # 表格数据
                if name == '资产负债率':
                    headers = [''] + [i[:4] for i in list(df_s.index)]
                    row1 = ['企业'] + [str(i) + '%' for i in y]
                    row2 = ['行业'] + [str(i) + '%' for i in y_ind]
                    row3 = ['差额'] + [str(k) + '%' for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                    row4 = ['差额百分比'] + [
                        str(round(k * 100, 2)) + '%' if abs(k * 100) < 50 else '*' + str(round(k * 100, 2)) + '%' for k
                        in
                        [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                    formatter_line = JsCode(
                        """function(params){
                        color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                        form = params[0].name + '年' + '<br />';
                        for(i=0; i<params.length; i++){
                        form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '%' + '<br />';
                        }
                        return form;
                        }
                        """
                    )
                else:
                    headers = [''] + [i[:4] for i in list(df_s.index)]
                    row1 = ['企业'] + [str(i) for i in y]
                    row2 = ['行业'] + [str(i) for i in y_ind]
                    row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                    row4 = ['差额百分比'] + [
                        str(round(k * 100, 2)) + '%' if abs(k * 100) < 50 else '*' + str(round(k * 100, 2)) + '%' for k
                        in
                        [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                    formatter_line = JsCode(
                        """function(params){
                        color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                        form = params[0].name + '年' + '<br />';
                        for(i=0; i<params.length; i++){
                        form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '<br />';
                        }
                        return form;
                        }
                        """
                    )

                line = (
                    Line(init_opts=opts.InitOpts(chart_id='id_' + str(count)))
                        .add_xaxis(xaxis_data=x)
                        .add_yaxis(series_name='企业',
                                   y_axis=y,
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[1]),
                                   is_connect_nones=True)
                        .add_yaxis(series_name='行业',
                                   y_axis=y_ind,
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[2]),
                                   is_connect_nones=True)
                        .add_yaxis(series_name='差额(企业-行业)',
                                   y_axis=[round(y[i] - y_ind[i], 2) for i in range(len(y))],
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[3]),
                                   is_connect_nones=True,
                                   label_opts=opts.LabelOpts(color=temp_color_lst[3])
                                   )
                        # .set_series_opts(areastyle_opts=opts.AreaStyleOpts(opacity=0.65))
                        .set_series_opts(linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                        .set_global_opts(
                        title_opts=opts.TitleOpts(title=name),
                        datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                        xaxis_opts=opts.AxisOpts(name='年份',
                                                 name_location='middle',
                                                 name_gap=25,
                                                 boundary_gap=False),
                        yaxis_opts=opts.AxisOpts(name='数\n值',
                                                 name_location='middle',
                                                 name_gap=40,
                                                 name_rotate=0,
                                                 splitline_opts=opts.SplitLineOpts(is_show=True)),
                        tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                      formatter=formatter_line,
                                                      axis_pointer_type='cross'),
                        legend_opts=opts.LegendOpts(type_='scroll', pos_left='20%', pos_right='20%'),
                        toolbox_opts=opts.ToolboxOpts(
                            feature=opts.ToolBoxFeatureOpts(
                                data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                                brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                                magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                            )
                        )
                    )
                )

                table = (
                    Table().add(headers=headers, rows=[row1, row2, row3, row4])
                )

                if name == '资产负债率':
                    line.set_series_opts(label_opts=opts.LabelOpts(is_show=True,
                                                                   position='top',
                                                                   font_size=12,
                                                                   formatter=formatter_line_label
                                                                   )
                                         )

                if name.find('资产负债率') >= 0:
                    dic_extra = DataProcess(self.ticker).get_extra_trial_data()
                    level_line = Line().add_xaxis(x)
                    for level in range(1, 6):
                        level_line.add_yaxis(level_val[level],
                                             dic_extra['DebtToAssetratio'][str(level)],
                                             is_symbol_show=False,
                                             label_opts=opts.LabelOpts(is_show=False),
                                             areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
                                             itemstyle_opts=opts.ItemStyleOpts(color=level_color[level])
                                             )
                    line.overlap(level_line)
                elif name.find('已获利息倍数') >= 0:
                    dic_extra = DataProcess(self.ticker).get_extra_trial_data()
                    level_line = Line().add_xaxis(x)
                    for level in range(1, 6):
                        level_line.add_yaxis(level_val[level],
                                             dic_extra['InterestCoverageRatio'][str(level)],
                                             is_symbol_show=False,
                                             label_opts=opts.LabelOpts(is_show=False),
                                             areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
                                             itemstyle_opts=opts.ItemStyleOpts(color=level_color[level])
                                             )
                    line.overlap(level_line)

                c_dic[name] = [line, table]
                count += 1

            # part3
            df_p = DataProcess(self.ticker).profitability()[0]
            df_p_ind = DataProcess(self.ticker).profitability()[1]
            rename_dic = dict(zip(df_p_ind.columns, df_p.columns))
            df_p_ind.rename(columns=rename_dic, inplace=True)
            x = [i[:4] for i in list(df_p.index)]

            for name in ['总资产净利率ROA', '净资产收益率ROE']:
                # 折线图数据
                y = [round(i, 2) for i in list(df_p[name])]
                y_ind = [round(i, 2) for i in list(df_p_ind[name])]

                # 表格数据
                if name == '每股收益EPS' or name == '每股营业总收入':
                    headers = ['单位:元'] + [i[:4] for i in list(df_p.index)]
                    row1 = ['企业'] + [str(i) for i in y]
                    row2 = ['行业'] + [str(i) for i in y_ind]
                    row3 = ['差额'] + [str(k) for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                    row4 = ['差额百分比'] + [
                        str(round(k * 100, 2)) + '%' if abs(k * 100) < 50 else '*' + str(round(k * 100, 2)) + '%' for k
                        in
                        [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                    subtitle = '单位: 元'
                    formatter_yaxis = '{value}'

                    formatter_line_label = JsCode(
                        """function(params) {
                        form = params.value[1] + '元';
                        return form;
                    }"""
                    )
                    formatter_line = JsCode(
                        """function(params){
                        color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                        form = params[0].name + '年' + '<br />';
                        for(i=0; i<params.length; i++){
                        form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '元' + '<br />';
                        }
                        return form;
                        }
                        """
                    )

                else:
                    headers = [''] + [i[:4] for i in list(df_p.index)]
                    row1 = ['企业'] + [str(i) + '%' for i in y]
                    row2 = ['行业'] + [str(i) + '%' for i in y_ind]
                    row3 = ['差额'] + [str(k) + '%' for k in [round(y[i] - y_ind[i], 2) for i in range(len(y))]]
                    row4 = ['差额百分比'] + [
                        str(round(k * 100, 2)) + '%' if abs(k * 100) < 50 else '*' + str(round(k * 100, 2)) + '%' for k
                        in
                        [((y[i] - y_ind[i]) / abs(y_ind[i])) for i in range(len(y))]]

                    subtitle = ''
                    formatter_yaxis = '{value}%'

                    formatter_line_label = JsCode(
                        """function(params) {
                        form = params.value[1] + '%';
                        return form;
                    }"""
                    )
                    formatter_line = JsCode(
                        """function(params){
                        color_icon = '<span style="display:inline-block;margin-right:5px;border-radius:50%;width:10px;height:10px;left:5px;background-color:';
                        form = params[0].name + '年' + '<br />';
                        for(i=0; i<params.length; i++){
                        form += color_icon + params[i].color + '"></span>'+ params[i].seriesName + '：' + params[i].data[1] + '%' + '<br />';
                        }
                        return form;
                        }
                        """
                    )

                line = (
                    Line(init_opts=opts.InitOpts(chart_id=name))
                        .add_xaxis(xaxis_data=x)
                        .add_yaxis(series_name='企业',
                                   y_axis=y,
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[1]),
                                   is_connect_nones=True)
                        .add_yaxis(series_name='行业',
                                   y_axis=y_ind,
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[2]),
                                   is_connect_nones=True)
                        .add_yaxis(series_name='差额(企业-行业)',
                                   y_axis=[round(y[i] - y_ind[i], 2) for i in range(len(y))],
                                   itemstyle_opts=opts.ItemStyleOpts(color=temp_color_lst[3]),
                                   is_connect_nones=True)
                        .set_series_opts(label_opts=opts.LabelOpts(formatter=formatter_line_label),
                                         linestyle_opts=opts.LineStyleOpts(width=WIDTH))
                        .set_global_opts(
                        title_opts=opts.TitleOpts(title=name, subtitle=subtitle),
                        datazoom_opts=[opts.DataZoomOpts(range_start=0, range_end=100, pos_bottom='-7px')],
                        xaxis_opts=opts.AxisOpts(name='年份',
                                                 name_location='middle',
                                                 name_gap=25,
                                                 boundary_gap=False),
                        yaxis_opts=opts.AxisOpts(name='数\n值',
                                                 name_location='middle',
                                                 name_gap=40,
                                                 name_rotate=0,
                                                 axislabel_opts=opts.LabelOpts(formatter=formatter_yaxis),
                                                 splitline_opts=opts.SplitLineOpts(is_show=True)),
                        tooltip_opts=opts.TooltipOpts(trigger='axis',
                                                      formatter=formatter_line,
                                                      axis_pointer_type='cross'),
                        legend_opts=opts.LegendOpts(type_='scroll', pos_left='20%', pos_right='20%'),
                        toolbox_opts=opts.ToolboxOpts(
                            feature=opts.ToolBoxFeatureOpts(
                                data_zoom=opts.ToolBoxFeatureDataZoomOpts(is_show=False),
                                brush=opts.ToolBoxFeatureBrushOpts(type_='clear'),
                                magic_type=opts.ToolBoxFeatureMagicTypeOpts(type_=['bar'])
                            )
                        )
                    )
                )

                table = (
                    Table().add(headers=headers,
                                rows=[row1, row2, row3, row4])
                )

                if name.find('ROE') >= 0:
                    dic_extra = DataProcess(self.ticker).get_extra_trial_data()
                    level_line = Line().add_xaxis(x)
                    for level in range(1, 6):
                        level_line.add_yaxis(level_val[level],
                                             dic_extra['ROE'][str(level)],
                                             is_symbol_show=False,
                                             label_opts=opts.LabelOpts(is_show=False),
                                             areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
                                             itemstyle_opts=opts.ItemStyleOpts(color=level_color[level])
                                             )
                    line.overlap(level_line)
                elif name.find('ROA') >= 0:
                    dic_extra = DataProcess(self.ticker).get_extra_trial_data()
                    level_line = Line().add_xaxis(x)
                    for level in range(1, 6):
                        level_line.add_yaxis(level_val[level],
                                             dic_extra['ROA'][str(level)],
                                             is_symbol_show=False,
                                             label_opts=opts.LabelOpts(is_show=False),
                                             areastyle_opts=opts.AreaStyleOpts(opacity=0.5),
                                             itemstyle_opts=opts.ItemStyleOpts(color=level_color[level])
                                             )
                    line.overlap(level_line)

                c_dic[name] = [line, table]

        else:
            logger.warning('该企业不是示例')

        return c_dic

    @staticmethod
    def paste_id() -> NoReturn:
        """
        寻找html中Table的id（因不是标准echarts图表，所以无法指定id）
        :return: NoReturn
        """
        soup = BeautifulSoup(open(r'.\data\raw_figure_html\Extra_Trial_raw.html', encoding='utf-8'), features="lxml")
        div_lst = soup.find_all('div', class_='chart-container')
        id_lst = []

        # 获得html的id列表
        for i in div_lst:
            id = re.search("(?<=id=\").+?(?=\")", str(i)).group()
            id_lst.append(id)
        # 寻找Table的id
        table_id_lst = []
        for id in id_lst:
            if len(id) > 10:
                table_id_lst.append(id)
            else:
                pass
        # 替换Table的id
        # 读
        with open(r'.\data\config_data\chart_config_ExtraPlot.json', 'rb') as r:
            count = 0
            params = json.load(r)
            for dic in params:
                if len(dic['cid']) > 10:
                    dic['cid'] = table_id_lst[count]
                    count += 1
                else:
                    pass
        r.close()
        # 写
        with open(r'.\data\config_data\chart_config_ExtraPlot.json', 'w') as w:
            json.dump(params, w)
        w.close()

    def plot(self):
        """
        整合图表，指定预设config
        :return: NoReturn
        """
        page = Page(page_title="Extra_Trial", layout=Page.DraggablePageLayout)
        for name in ['应收账款周转率', '总资产周转率', '资产负债率', '已获利息倍数', '总资产净利率ROA', '净资产收益率ROE']:  # '存货周转率'
            page.add(self.trial_figures()[name][0]).add(self.trial_figures()[name][1])
        raw_path = r'.\data\raw_figure_html\Extra_Trial_raw.html'
        cfg_path = r'.\data\config_data\chart_config_ExtraPlot.json'
        ripe_path = r'.\data\figure_html\Extra_Trial.html'
        page.render(raw_path)
        ExtraPlot(self.ticker).paste_id()
        Page.save_resize_html(source=raw_path,
                              cfg_file=cfg_path,
                              dest=ripe_path)
        webbrowser.open(ripe_path)


if __name__ == '__main__':
    # page = StatementOfProfitAndLoss()
    # page = StatementOfFinancialPostition()
    # page.plot()
    # webbrowser.open('render.html')

    # ticker = '002069.SZ'
    # c = PlotAll(ticker)
    # c.plot()

    # ticker_lst = ['600248.SZ', '601369.SH', '688596.SH', '000540.SZ', '002499.SZ',
    #               '300757.SZ', '300345.SZ', '300133.SZ', '002288.SZ', '300588.SZ',
    #               '000676.SZ', '300273.SZ', '300485.SZ', '002089.SZ', '002647.SZ',
    #               '002798.SZ', '003012.SZ', '601137.SH', '300336.SZ', '600545.SH',
    #               '002485.SZ', '300356.SZ', '002569.SZ', '000820.SZ', '600310.SH',
    #               '600666.SH', '600712.SH', '000687.SH', '300173.SZ', '000525.SZ',
    #               '000663.SZ', '002164.SZ', '603613.SH', '300518.SZ', '002127.SZ',
    #               '002024.SZ', '002280.SZ', '300792.SZ', '000002.SZ', '000006.SZ',
    #               '000007.SZ', '000009.SZ', '000010.SZ', '000011.SZ', '000012.SZ',
    #               '000014.SZ', '000016.SZ', '000017.SZ', '000019.SZ', '000020.SZ']

    df = pd.read_excel(r'D:\Programming\Python\Code\sophomore_year\KPMG\data\trial_data\id.xlsx', index_col=0)
    ticker_lst = df.iloc[:, 0]
    ticker_lst = [str(z).zfill(6) for z in ticker_lst]
    ticker_lst = [z + '.SH' if z[:1] == '6' else z + '.SZ' for z in ticker_lst]

    # ticker_lst = [z + '.SH' if z[:1] == '6' else z + '.SZ' for z in s]
    # ticker_lst = [z.strip() for z in s]
    # ticker_lst = [z + '.SH' for z in ticker_lst if z[:1] == '6']

    ticker_lst = ['000509.SZ']

    for ticker in ticker_lst:
        c = PlotAll(ticker)
        c.plot()
        # c = DataProcess(ticker)
        # c.excel_analysis2_judge_law
