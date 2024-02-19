# -*- coding:utf-8 -*-
"""
 作者: QL
 日期: 2022年09月03日
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
import pandas as pd
import numpy as np


class DataProcess:
    def __init__(self, ticker):
        self.ticker = ticker

    def excel_analysis1(self) -> list:
        """
        获得、处理财务分析数据1
        :return: 财务分析表格DataFrame列表[企业, 行业]
        """
        # 企业
        path1 = r".\\data\\" + self.ticker + r"_舞弊分析数据.csv"
        df1 = pd.read_csv(path1, index_col=0)

        # 数据处理
        # 增长率计算
        growth_term = "应收账款 应收票据 其他应收款合计 预付款项 存货 在建工程 长期待摊费用 固定资产 " \
                      "商誉 资产减值准备 短期借款 应付票据 应付账款 其他应付款合计 " \
                      "营业总收入 主营营业收入 营业总成本 销售毛利率 销售净利率 " \
                      "经营活动产生的现金流量净额".split(' ')
        for term in growth_term:
            df1[term + '增长率'] = (df1[term] - df1[term].shift(1)) / df1[term].shift(1)

        for term in ['应收账款', '存货', '固定资产']:
            df1[term + '比率'] = df1[term] / df1['负债和股东权益合计']

        df1['费用率'] = (df1['营业总成本'] + df1['销售费用'] + df1['管理费用'] + df1['财务费用']) / df1['营业总收入']
        df1['研发支出占比'] = df1['研发费用'] / df1['营业总成本']
        df1['带息负债和货币资金比值'] = df1['带息负债'] / df1['货币资金']

        # sorted_term = [term + '增长率' for term in growth_term[: -3]] + \
        #               ['总资产周转率', '净资产收益率'] + \
        #               [term + '增长率' for term in growth_term[-3: -1]] + ['销售毛利率', '销售净利率', '费用率'] + \
        #               ['存货增长率', '固定资产增长率', '应收账款比率', '应收账款周转率', '存货比率', '存货周转率', '固定资产比率', '固定资产周转率'] + \
        #               ['研发支出占比', '经营活动产生的现金流量净额增长率', '资产负债率', '带息负债和货币资金比值']
        # return df.iloc[1:, :].loc[:, sorted_term].T

        # 行业
        path2_1 = r".\\data\\" + self.ticker + r"_舞弊分析数据_行业1.csv"
        path2_2 = r".\\data\\" + self.ticker + r"_舞弊分析数据_行业2.csv"
        df2_1 = pd.read_csv(path2_1, index_col=0)
        df2_2 = pd.read_csv(path2_2, index_col=0)
        df2 = pd.concat([df2_1, df2_2], axis=1)

        # 数据处理
        # 增长率计算
        growth_term_hy = "应收账款 预付款项 存货 固定资产 应付账款 营业总收入 营业总成本 销售毛利率 销售净利率 经营活动产生的现金流量净额".split(' ')
        for term in growth_term_hy:
            df2[term + '增长率'] = (df1[term] - df1[term].shift(1)) / df1[term].shift(1)

        for term in ['应收账款', '存货', '固定资产']:
            df2[term + '比率'] = df2[term] / df2['负债和股东权益合计']

        df2['费用率'] = (df2['营业总成本'] + df2['销售费用'] + df2['管理费用'] + df2['财务费用']) / df2['营业总收入']

        return [df1.iloc[1:, :], df2.iloc[1:, :]]

    def excel_analysis2(self) -> list:
        """
        获得、处理财务分析数据2
        :return: 财务分析表格DataFrame
        """
        path = r".\\data\\" + self.ticker + r"_舞弊分析数据.csv"
        df = pd.read_csv(path, index_col=0)

        # 数据计算
        def groth_rate(df, col):
            return (df[col] - df[col].shift(1)) / df[col].shift(1)

        # 固定资产占总资产的比重变化率
        df['固定资产占总资产的比重'] = df['固定资产'] / df['负债和股东权益合计']
        df['固定资产占总资产的比重变化率'] = groth_rate(df, '固定资产占总资产的比重')

        # 固定资产增加率
        df['固定资产增加率'] = groth_rate(df, '固定资产')
        # 累计折旧占固定资产原值的比重变化率
        df['累计折旧占固定资产原值的比重'] = df['固定资产累计折旧'] / df['固定资产']
        df['累计折旧占固定资产原值的比重变化率'] = groth_rate(df, '累计折旧占固定资产原值的比重')

        # 存货占总资产比重变化率
        df['存货占总资产比重'] = df['存货'] / df['负债和股东权益合计']
        df['存货占总资产比重变化率'] = groth_rate(df, '存货占总资产比重')

        # 存货增加率
        df['存货增加率'] = groth_rate(df, '存货')
        # 营业收入增长率
        df['营业收入增长率'] = groth_rate(df, '营业总收入')
        # 营业成本增长率
        df['营业成本增长率'] = groth_rate(df, '营业总成本')

        # 存货跌价准备占存货的比重变化率
        df['存货跌价准备占存货的比重'] = df['存货跌价准备合计'] / df['应收账款合计']
        df['存货跌价准备占存货的比重变化率'] = groth_rate(df, '存货跌价准备占存货的比重')

        # 应收账款增长率
        df['应收账款增长率'] = groth_rate(df, '应收账款')
        # 存货增长率
        df['存货增长率'] = groth_rate(df, '存货')
        # 营业收入增长率
        df['营业收入增长率'] = groth_rate(df, '营业总收入')

        # 现金及现金等价物占总资产的比重变化率
        df['现金及现金等价物占总资产的比重'] = df['期末现金及现金等价物余额'] / df['负债和股东权益合计']
        df['现金及现金等价物占总资产的比重变化率'] = groth_rate(df, '现金及现金等价物占总资产的比重')
        # 应收账款占总资产的比重增长率
        df['应收账款占总资产的比重'] = df['应收账款'] / df['负债和股东权益合计']
        df['应收账款占总资产的比重增长率'] = groth_rate(df, '应收账款占总资产的比重')

        # 存货占总资产比重变化率
        df['存货占总资产比重'] = df['存货'] / df['负债和股东权益合计']
        df['存货占总资产比重变化率'] = groth_rate(df, '存货占总资产比重')

        # 坏账占应收账款比重的变化率
        df['坏账占应收账款比重'] = df['坏账准备合计'] / df['应收账款合计']
        df['坏账占应收账款比重变化率'] = groth_rate(df, '坏账占应收账款比重')

        # 销售费用增长速度
        df['销售费用增长速度'] = groth_rate(df, '销售费用')
        # 营业收入增长速度
        df['营业收入增长速度'] = groth_rate(df, '营业总收入')

        # 毛利率变化率
        df['毛利率变化率'] = groth_rate(df, '销售毛利率')

        # 毛利率变化率
        # 存货周转率变化率
        df['存货周转率变化率'] = groth_rate(df, '存货周转率')

        # 毛利率变化率
        # 应收账款周转率变化率
        df['应收账款周转率变化率'] = groth_rate(df, '应收账款周转率')

        # 毛利率变化率
        # 应付账款变化率
        df['应付账款变化率'] = groth_rate(df, '应付账款')

        # 毛利率变化率
        # 现金循环周期
        df['现金循环周期'] = df['存货周转天数'] + df['应收账款周转天数'] - df['应付账款周转天数']

        return df.iloc[1:, :]


class ExcelAnalysis1:
    def __init__(self, ticker):
        self.ticker = ticker

    def write(self):
        # 企业数据
        df1 = DataProcess(self.ticker).excel_analysis1()[0]  # 读取数据
        df1.drop(labels=['CODES'], axis=1, inplace=True)  # 删除第一列（股票代码）
        df1.index = [z[: 4] for z in df1.index]  # 时间标签年化（2021-12-31 -> 2021）
        df1.rename(mapper=lambda x: str(x) + '_企业', axis=0, inplace=True)  # 年份标签重命名（2021 -> 2021_企业）

        # 行业数据
        df2 = DataProcess(self.ticker).excel_analysis1()[1]
        df2.drop(labels=['CODES'], axis=1, inplace=True)
        df2.index = [z[: 4] for z in df2.index]
        df2.rename(mapper=lambda x: str(x) + '_行业', axis=0, inplace=True)

        path = r'./data/excel_template/excel_eco_template.xlsx'
        wb = load_workbook(path)

        # 企业数据写入模板 #
        ws1 = wb[wb.sheetnames[0]]

        # 年份修改
        for i, col in enumerate(['B', 'C', 'D', 'E', 'F']):
            coordinate = col + '1'
            ws1[coordinate].value = df1.index.tolist()[i]

        data_rows_index = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 19, 20, 21, 22, 23, 25, 26,
                           27, 28, 29, 31, 32, 33, 34, 35, 36, 37, 38, 40, 41, 42, 43]  # 需要写入数据的行
        data_rows_name = ['应收账款增长率', '应收票据增长率', '其他应收款合计增长率', '预付款项增长率', '存货增长率',
                          '在建工程增长率', '长期待摊费用增长率', '固定资产增长率', '商誉增长率', '资产减值准备增长率',
                          '短期借款增长率', '应付票据增长率', '应付账款增长率', '其他应付款合计增长率',
                          '营业总收入增长率', '主营营业收入增长率', '营业总成本增长率', '总资产周转率', '净资产收益率',
                          '销售毛利率增长率', '销售净利率增长率', '销售毛利率', '销售净利率', '费用率',
                          '存货增长率', '固定资产增长率', '应收账款比率', '应收账款周转率', '存货比率', '存货周转率', '固定资产比率', '固定资产周转率',
                          '研发支出占比', '经营活动产生的现金流量净额增长率', '资产负债率', '带息负债和货币资金比值']
        data_rows = dict(zip(data_rows_index, data_rows_name))
        for row, name in data_rows.items():
            for i, col in enumerate(['B', 'C', 'D', 'E', 'F']):
                coordinate = col + str(row)
                cell = ws1[coordinate]
                cell.value = df1[name][i]

        # 企业vs行业数据写入模板 #
        ws2 = wb[wb.sheetnames[1]]

        # 年份修改
        for i, col in enumerate(['B', 'D', 'F', 'H', 'J']):
            coordinate = col + '1'
            ws2[coordinate].value = df1.index.tolist()[i]
        for j, col in enumerate(['C', 'E', 'G', 'I', 'K']):
            coordinate = col + '1'
            ws2[coordinate].value = df2.index.tolist()[i]

        data_rows_index_2 = [3, 4, 5, 6, 8, 10, 11, 12, 13, 15, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27, 28, 30, 31]
        data_rows_name_2 = ['应收账款增长率', '预付款项增长率', '存货增长率', '固定资产增长率', '应付账款增长率', '营业总收入增长率',
                            '营业总成本增长率', '总资产周转率', '净资产收益率', '销售毛利率增长率', '销售净利率增长率', '销售毛利率',
                            '销售净利率', '费用率', '存货增长率', '固定资产增长率', '应收账款比率', '应收账款周转率', '存货比率',
                            '存货周转率', '固定资产比率', '固定资产周转率', '经营活动产生的现金流量净额增长率', '资产负债率']
        data_rows_2 = dict(zip(data_rows_index_2, data_rows_name_2))
        for row, name in data_rows_2.items():
            for i, col in enumerate(['B', 'D', 'F', 'H', 'J']):
                coordinate = col + str(row)
                cell = ws2[coordinate]
                cell.value = df1[name][i]
            for j, col in enumerate(['C', 'E', 'G', 'I', 'K']):
                coordinate = col + str(row)
                cell = ws2[coordinate]
                cell.value = df2[name][j]

        save_path = r'.\data\excel_data\\' + self.ticker + '_舞弊财务指标异常风险分析.xlsx'
        wb.save(save_path)

    # def write(self):
    #     df1 = DataProcess('10').excel_analysis1()[0]
    #     df1.index = [z[: 4] for z in df1.index]
    #     df1.rename(mapper=lambda x: str(x) + '_企业', axis=0, inplace=True)
    #     df2 = DataProcess('10').excel_analysis1()[1]
    #     df2.index = [z[: 4] for z in df2.index]
    #     df2.rename(mapper=lambda x: str(x) + '_行业', axis=0, inplace=True)
    #     df_lst = []
    #     for i in range(df1.shape[1]):
    #         temp_df = pd.concat([df1.iloc[:, i], df2.iloc[:, i]], axis=1)
    #         df_lst.append(temp_df)
    #     df = pd.concat(df_lst, axis=1)
    #     df.to_excel("舞弊分析数据excel分析.xlsx")
    #
    # def revise(self):
    #     wb = load_workbook("舞弊新内容——表格生成/舞弊分析数据excel分析.xlsx")
    #     ws = wb[wb.sheetnames[0]]
    #
    #     # 自适应列宽
    #     def col_width_autofit(sheet):
    #         n_col = sheet.max_column
    #         n_row = sheet.max_row
    #
    #         col_num_dict = dict(zip(np.arange(1, 12), 'A,B,C,D,E,F,G,H,I,J,K'.split(',')))
    #         for i in range(1, n_col + 1):
    #             max_width = 0
    #             for j in range(1, n_row + 1):
    #                 iter_width = 0
    #                 sheet_value = sheet.cell(column=i, row=j).value
    #
    #                 if sheet_value is None:
    #                     pass
    #                 else:
    #                     for item in str(sheet_value):
    #                         # if item.isdigit() or item.isalpha():  # isalpha包含了中文字符
    #                         if item.isdigit() or ('A' <= item <= 'Z' or 'a' <= item <= 'z'):
    #                             iter_width += 0.7
    #                         else:
    #                             iter_width += 2.2
    #
    #                 if max_width < iter_width:
    #                     max_width = iter_width
    #
    #             sheet.column_dimensions[col_num_dict[i]].width = max_width
    #
    #     # 插入区块标题
    #     title_style = NamedStyle(name='title_style',
    #                              font=Font(color=Color('DC143C')),
    #                              fill=PatternFill(patternType='solid', fgColor=Color('FFFF00')),
    #                              border=Border(left=Side(style='thin', color=Color('000000')),
    #                                            right=Side(style='thin', color=Color('000000')),
    #                                            top=Side(style='thin', color=Color('000000')),
    #                                            bottom=Side(style='thin', color=Color('000000'))),
    #                              alignment=Alignment(horizontal='center', vertical='center'))
    #
    #     ws.insert_rows(2)
    #     ws.cell(column=1, row=2, value='资产类').style = title_style
    #     ws.insert_rows(2 + 11)
    #     ws.cell(column=1, row=2 + 11, value='负债类').style = title_style
    #     ws.insert_rows(2 + 11 + 5)
    #     ws.cell(column=1, row=2 + 11 + 5, value='资产带来收入').style = title_style
    #     ws.insert_rows(2 + 11 + 5 + 6)
    #     ws.cell(column=1, row=2 + 11 + 5 + 6, value='收入带来利润').style = title_style
    #     ws.insert_rows(2 + 11 + 5 + 6 + 6)
    #     ws.cell(column=1, row=2 + 11 + 5 + 6 + 6, value='利润带来现金流').style = title_style
    #     ws.insert_rows(2 + 11 + 5 + 6 + 6 + 9)
    #     ws.cell(column=1, row=2 + 11 + 5 + 6 + 6 + 9, value='其他').style = title_style
    #
    #     col_width_autofit(ws)
    #
    #     wb.save('舞弊分析数据excel分析.xlsx')

    def analysis(self):
        self.write()
        self.revise()


class ExcelAnalysis2:
    def __init__(self, ticker):
        self.ticker = ticker

    def write_calculation_table(self):
        df = DataProcess(self.ticker).excel_analysis2()  # 读取数据
        df.drop(labels=['CODES'], axis=1, inplace=True)  # 删除第一列（股票代码）
        df.index = [z[: 4] for z in df.index]  # 时间标签年化（2021-12-31 -> 2021）

        path = r'./data/excel_template/excel_template.xlsx'
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]

        # 年份修改
        for i, col in enumerate(['B', 'C', 'D', 'E', 'F']):
            coordinate = col + '1'
            ws[coordinate].value = df.index.tolist()[i]

        data_rows_index = [4, 7, 8, 12, 15, 16, 17, 20, 23, 24, 25, 29, 30, 32, 34, 39, 40, 42, 44, 45, 47, 48, 50, 51, 53, 54]  # 需要写入数据的行
        data_rows_name = ['固定资产占总资产的比重变化率', '固定资产增加率', '累计折旧占固定资产原值的比重变化率', '存货占总资产比重变化率',
                          '存货增加率', '营业收入增长率', '营业成本增长率', '存货跌价准备占存货的比重变化率', '应收账款增长率', '存货增长率',
                          '营业收入增长率', '现金及现金等价物占总资产的比重变化率', '应收账款占总资产的比重增长率', '存货占总资产比重变化率',
                          '坏账占应收账款比重变化率', '销售费用增长速度', '营业收入增长速度', '毛利率变化率', '毛利率变化率', '存货周转率变化率',
                          '毛利率变化率', '应收账款周转率变化率', '毛利率变化率', '应付账款变化率', '毛利率变化率', '现金循环周期']
        data_rows = dict(zip(data_rows_index, data_rows_name))
        for row, name in data_rows.items():
            for i, col in enumerate(['B', 'C', 'D', 'E', 'F']):
                coordinate = col + str(row)
                cell = ws[coordinate]
                cell.value = df[name][i]

        wb.save(path)

    def write_exhibition_table(self):
        pass

    def write_all(self):
        self.write_calculation_table()


if __name__ == '__main__':
    ticker = '000002.SZ'
    # df1 = DataProcess(ticker).excel_analysis1()[0]
    # df2 = DataProcess(ticker).excel_analysis1()[1]
    # print(df1)
    # print(df2)

    ExcelAnalysis1(ticker).write()

    # a = ExcelAnalysis('10')
    # a.write()
